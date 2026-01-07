import streamlit as st
import requests
import json
import pandas as pd
import time
import os
import datetime
from utils import (
    get_current_base_url, 
    get_user_specific_paths,
    cookies_dict_to_string,
    cookies_string_to_dict,
    load_user_apis,
    save_user_apis,
    load_api_configs,
    save_api_config,
    load_api_history,
    save_api_history,
    load_cookies_config,
    save_cookies_config,
    make_http_request,
    get_response_content,
    create_history_entry,
    get_existing_users,
    validate_username,
    ensure_path_format,
    load_environments_config,
    save_environments_config,
    get_enabled_environments
)


# Global admin cookies file
ADMIN_COOKIES_FILE = os.path.join(os.path.dirname(__file__), "admin_cookies_config.json")

# Global API configurations file
API_CONFIGS_FILE = os.path.join(os.path.dirname(__file__), "api_configs.json")


def load_help_content():
    """Load help content from markdown file"""
    try:
        help_file_path = os.path.join(os.path.dirname(__file__), "HUONG_DAN_SU_DUNG.md")
        if os.path.exists(help_file_path):
            with open(help_file_path, 'r', encoding='utf-8') as f:
                return f.read()
        else:
            return "Help file not found."
    except Exception as e:
        return f"Error reading help file: {str(e)}"


def load_admin_content():
    """Load admin-editable content from markdown file"""
    try:
        admin_content_file = os.path.join(os.path.dirname(__file__), "ADMIN_CONTENT.md")
        if os.path.exists(admin_content_file):
            with open(admin_content_file, 'r', encoding='utf-8') as f:
                return f.read()
        else:
            # Default content if file doesn't exist
            return "# Admin Editable Content\n\nThis content can be edited by administrators.\n\n## Welcome\n\nWelcome to the API Client Tester!\n\n## Getting Started\n\nPlease contact your administrator for more information."
    except Exception as e:
        return f"Error reading admin content file: {str(e)}"


def save_admin_content(content):
    """Save admin-editable content to markdown file"""
    try:
        admin_content_file = os.path.join(os.path.dirname(__file__), "ADMIN_CONTENT.md")
        with open(admin_content_file, 'w', encoding='utf-8') as f:
            f.write(content)
        return True
    except Exception as e:
        st.error(f"Error saving admin content: {str(e)}")
        return False


def _update_admin_content_with_timer_jobs():
    """Auto-update ADMIN_CONTENT.md with timer job URLs (SIT only)"""
    try:
        # Get current timer job entries
        timer_entries = st.session_state.get('timer_job_entries', [])
        
        if not timer_entries:
            # If no timer entries, just ensure ADMIN_CONTENT.md exists with basic content
            current_content = load_admin_content()
            if "No timer job entries" not in current_content:
                # Remove any existing timer job sections and add empty notice
                lines = current_content.split('\n')
                filtered_lines = []
                skip_section = False
                
                for line in lines:
                    if line.startswith('## ') and ('Timer' in line or 'timer' in line):
                        skip_section = True
                        continue
                    elif line.startswith('## ') and skip_section:
                        skip_section = False
                        filtered_lines.append(line)
                    elif not skip_section and not line.startswith('```'):
                        filtered_lines.append(line)
                    elif not skip_section and line.startswith('```') and len(filtered_lines) > 0 and not filtered_lines[-1].startswith('```'):
                        # Skip code blocks that were part of timer sections
                        continue
                    elif not skip_section:
                        filtered_lines.append(line)
                
                # Clean up and add basic content if needed
                clean_content = '\n'.join(filtered_lines).strip()
                if not clean_content:
                    clean_content = "# Admin Content\n\nWelcome to the API Tester system.\n\n*No timer job entries configured yet.*"
                
                save_admin_content(clean_content)
            return True
        
        # Check if SIT environment exists
        enabled_envs = get_enabled_environments()
        if "SIT" not in enabled_envs:
            return False
        
        # Load current admin content
        current_content = load_admin_content()
        
        # Remove existing timer job sections
        lines = current_content.split('\n')
        filtered_lines = []
        skip_section = False
        
        for line in lines:
            if line.startswith('## ') and ('Timer' in line or 'timer' in line):
                skip_section = True
                continue
            elif line.startswith('## ') and skip_section:
                skip_section = False
                filtered_lines.append(line)
            elif not skip_section and not line.startswith('```'):
                filtered_lines.append(line)
            elif not skip_section and line.startswith('```') and len(filtered_lines) > 0 and not filtered_lines[-1].startswith('```'):
                # Skip code blocks that were part of timer sections
                continue
            elif not skip_section:
                filtered_lines.append(line)
        
        # Build new content with timer jobs
        base_content = '\n'.join(filtered_lines).strip()
        
        # Add timer job sections
        timer_content = ""
        for entry in timer_entries:
            # Get SIT URL
            base_url = get_current_base_url("SIT", entry['module'])
            final_path = entry['path'].replace('{timer_job_id}', entry['id'])
            full_url = f"{base_url}{final_path}"
            
            # Add timer job section with the specified format
            timer_content += f"\n\n## {entry['name']}:\n```\n{full_url}\n```"
        
        # Combine content
        if base_content:
            new_content = base_content + timer_content
        else:
            new_content = "# Timer Job APIs" + timer_content
        
        # Save updated content
        return save_admin_content(new_content)
        
    except Exception as e:
        print(f"Error updating admin content with timer jobs: {str(e)}")
        return False


def generate_timer_job_markdown():
    """Generate markdown content for Timer Job APIs from ADMIN_CONTENT.md"""
    try:
        # Load admin content which contains timer job URLs
        admin_content = load_admin_content()
        
        # Check if there are timer job entries in the content
        if "##" in admin_content and ("Timer" in admin_content or "timer" in admin_content or "```" in admin_content):
            # Return the admin content directly as it already contains formatted timer jobs
            return admin_content
        else:
            # No timer jobs found in admin content
            markdown_content = "# Timer Job APIs\n\n"
            markdown_content += "❌ **No Timer Job APIs configured yet.**\n\n"
            markdown_content += "💡 **Tip:** Admin users can add Timer Job APIs in the Admin Panel → Timer Run tab.\n\n"
            markdown_content += "📝 **Note:** Timer Job URLs are automatically generated for SIT environment only.\n"
            return markdown_content
        
    except Exception as e:
        return f"# Timer Job APIs\n\n❌ **Error loading Timer Job information:** {str(e)}\n"


@st.dialog("📖 Help", width="large")
def show_help_dialog():
    """Show help dialog with markdown content"""
    help_content = load_help_content()
    st.markdown(help_content)


@st.dialog("📝 Admin Content", width="large")
def show_admin_content_dialog():
    """Show admin content dialog with view/edit functionality"""
    # Check if user is admin (for editing)
    is_admin = st.session_state.get("is_admin", False)
    
    # Load current content
    admin_content = load_admin_content()
    
    if is_admin:
        # Admin can edit content
        st.subheader("🔧 Admin Mode - Edit Content")
        st.info("You can edit the content below. Changes will be saved immediately when you click Save.")
        
        # Text area for editing
        edited_content = st.text_area(
            "Edit Content (Markdown Format)",
            value=admin_content,
            height=400,
            help="Use Markdown syntax for formatting. Changes will be visible to all users."
        )
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("💾 Save Changes", type="primary"):
                if save_admin_content(edited_content):
                    st.success("Content saved successfully!")
                    st.rerun()
                else:
                    st.error("Failed to save content")
        
        with col2:
            if st.button("🔄 Reset to Default"):
                default_content = "# Admin Editable Content\n\nThis content can be edited by administrators.\n\n## Welcome\n\nWelcome to the API Client Tester!\n\n## Getting Started\n\nPlease contact your administrator for more information."
                if save_admin_content(default_content):
                    st.success("Content reset to default!")
                    st.rerun()
        
        st.markdown("---")
        st.subheader("📄 Preview")
        st.markdown(edited_content)
        
    else:
        # Regular users can only view content
        st.markdown(admin_content)


@st.dialog("⏰ Timer Job APIs", width="large")
def show_timer_job_dialog():
    """Show Timer Job APIs information in markdown format"""
    # Generate Timer Job markdown content
    timer_markdown = generate_timer_job_markdown()
    
    # Display the content
    st.markdown(timer_markdown)
    
    # Add refresh button for admins
    if st.session_state.get("is_admin", False):
        st.markdown("---")
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🔄 Refresh Timer Job Info"):
                st.rerun()
        
        with col2:
            if st.button("⚙️ Manage in Admin Panel"):
                st.session_state.admin_mode = True
                st.session_state.admin_tab_focus = "timer_run"
                st.rerun()


def save_to_predefined(api_name, api, file_paths):
    """Save API configuration to predefined configs shared by all users"""
    # Prepare config for saving
    save_config = api.copy()

    # Remove the full URL but keep the path
    if "url" in save_config:
        del save_config["url"]

    # Remove cookies as they are environment-specific
    if "cookies" in save_config:
        del save_config["cookies"]
    if "custom_cookies_string" in save_config:
        del save_config["custom_cookies_string"]

    # Save to predefined configurations
    configs = load_api_configs(file_paths["API_CONFIG_FILE"])
    configs[api_name] = save_config
    if save_api_config(configs, file_paths["API_CONFIG_FILE"]):
        st.success(f"API configuration for '{api_name}' saved to predefined configurations!")
    else:
        st.error("Failed to save to predefined configurations")


def show_multi_user_login():
    """Display the multi-user login page"""
    # Help and content management buttons above title
    col1, col2, col3 = st.columns([1, 1, 15])
    
    with col1:
        if st.button("📖 Instruction", key="login_help", help="Take a look about instruction"):
            show_help_dialog()
    
    with col2:
        if st.button("⏰ Timer Run", key="login_content", help="View Timer Job APIs information"):
            show_timer_job_dialog()
    
    # Title
    st.title("API Client Tester")

    # Initialize logged in users if not exists
    if 'logged_in_users' not in st.session_state:
        st.session_state.logged_in_users = {}
    if 'active_user' not in st.session_state:
        st.session_state.active_user = None

    # Show currently logged in users
    if st.session_state.logged_in_users:
        st.subheader("Currently Logged In Users")
        cols = st.columns(len(st.session_state.logged_in_users) + 1)
        
        # Sort users to prioritize QA and BA users
        qa_users = [(username, user_data) for username, user_data in st.session_state.logged_in_users.items() 
                   if username.upper().startswith("QA")]
        ba_users = [(username, user_data) for username, user_data in st.session_state.logged_in_users.items() 
                   if username.upper().startswith("BA")]
        regular_users = [(username, user_data) for username, user_data in st.session_state.logged_in_users.items() 
                        if not username.upper().startswith("QA") and not username.upper().startswith("BA")]
        
        # Sort each group alphabetically
        qa_users.sort(key=lambda x: x[0])
        ba_users.sort(key=lambda x: x[0])
        regular_users.sort(key=lambda x: x[0])
        
        # Combine with QA users first, then BA users, then regular users
        sorted_users = qa_users + ba_users + regular_users
        
        for i, (username, user_data) in enumerate(sorted_users):
            with cols[i]:
                is_active = st.session_state.active_user == username
                user_type = "Admin" if user_data.get('is_admin', False) else "User"
                status = "🟢 Active" if is_active else "⚫ Inactive"
                
                st.write(f"**{username}** ({user_type})")
                st.write(status)
                
                if not is_active:
                    if st.button(f"Switch to {username}", key=f"switch_{username}"):
                        st.session_state.active_user = username
                        _load_user_data(username)
                        st.rerun()
                
                if st.button(f"Logout {username}", key=f"logout_{username}"):
                    _logout_user(username)
                    st.rerun()

        # Add separator
        st.markdown("---")

        # Add New User button
        if st.button("➕ Add New User"):
            st.session_state.show_main_app = False
            st.rerun()

    # Get list of existing users
    user_dir = os.path.join(os.path.dirname(__file__), "user_data")
    if not os.path.exists(user_dir):
        os.makedirs(user_dir)

    # Get existing users but exclude already logged in users
    existing_users = get_existing_users()
    available_users = [u for u in existing_users if u not in st.session_state.logged_in_users]

    # Login options
    login_option = st.radio("Login Option", ["Existing User", "New User", "Admin Login"])

    if login_option == "Existing User":
        if available_users:
            username = st.selectbox("Select Username", available_users)
            login_button = st.button("Login")

            if login_button:
                _handle_multi_user_login(username, False)
        else:
            st.info("No available users to add. All existing users are already logged in.")

    elif login_option == "Admin Login":
        if "adminadmin" not in st.session_state.logged_in_users:
            st.subheader("Administrator Login")
            admin_password = st.text_input("Password", type="password", placeholder="Enter admin password")
            admin_login_button = st.button("Add Admin")

            if admin_login_button:
                if admin_password == "adminadmin":
                    _handle_multi_user_login("adminadmin", True)
                    st.success("Administrator added to session")
                    st.rerun()
                else:
                    st.error("Invalid admin password")
        else:
            st.info("Administrator is already logged in.")

    else:  # New User
        new_username = st.text_input("Enter New Username")
        create_button = st.button("Create and Add User")

        if create_button and new_username:
            all_users = list(st.session_state.logged_in_users.keys()) + existing_users
            # error_msg = validate_username(new_username, all_users)
            error_msg = ""
            if error_msg:
                st.error(error_msg)
            else:
                # Create user directory
                new_user_dir = os.path.join(user_dir, new_username)
                if not os.path.exists(new_user_dir):
                    os.makedirs(new_user_dir)

                _handle_multi_user_login(new_username, False)
                st.success(f"User {new_username} created and added to session")
                st.rerun()

    # Button to proceed to main app
    if st.session_state.logged_in_users:
        if st.button("🚀 Start Using API Tester", type="primary"):
            if not st.session_state.active_user:
                # Prioritize QA users, then BA users when selecting the first active user
                qa_users = [user for user in st.session_state.logged_in_users.keys() 
                           if user.upper().startswith("QA")]
                ba_users = [user for user in st.session_state.logged_in_users.keys() 
                           if user.upper().startswith("BA")]
                regular_users = [user for user in st.session_state.logged_in_users.keys() 
                                if not user.upper().startswith("QA") and not user.upper().startswith("BA")]
                
                # Sort each group and combine with QA users first, then BA users
                qa_users.sort()
                ba_users.sort()
                regular_users.sort()
                sorted_users = qa_users + ba_users + regular_users
                
                # Set first user as active (QA user if available, then BA user)
                st.session_state.active_user = sorted_users[0]
                _load_user_data(st.session_state.active_user)

            st.session_state.show_main_app = True
            st.rerun()


def _handle_multi_user_login(username: str, is_admin: bool):
    """Handle multi-user login setup"""
    # Get enabled environments
    enabled_envs = get_enabled_environments()
    
    # Determine default environment based on username
    if username.upper().startswith("BA") and "DEMO" in enabled_envs:
        default_env = "DEMO"
    else:
        default_env = "SIT" if "SIT" in enabled_envs else enabled_envs[0] if enabled_envs else "SIT"
    
    # Add user to logged in users
    st.session_state.logged_in_users[username] = {
        'is_admin': is_admin,
        'file_paths': get_user_specific_paths(username),
        'apis': {},
        'api_responses': {},
        'current_env': default_env,
        'api_history': [],
        'cookies_config': {}  # Start with empty cookies - will be loaded properly in _load_user_data
    }

    # Set as active user if it's the first one, if no active user, or if it's a QA/BA user
    if not st.session_state.active_user or len(st.session_state.logged_in_users) == 1:
        st.session_state.active_user = username
        _load_user_data(username)
    # If the new user is a QA user and the current active user is not, switch to the QA user
    elif username.upper().startswith("QA") and not st.session_state.active_user.upper().startswith("QA"):
        st.session_state.active_user = username
        _load_user_data(username)
    # If the new user is a BA user and the current active user is not QA or BA, switch to the BA user
    elif (username.upper().startswith("BA") and 
          not st.session_state.active_user.upper().startswith("QA") and 
          not st.session_state.active_user.upper().startswith("BA")):
        st.session_state.active_user = username
        _load_user_data(username)

    if not is_admin:
        st.success(f"User {username} added to session")


def _load_user_data(username: str):
    """Load data for the specified user and set as active"""
    if username not in st.session_state.logged_in_users:
        return

    user_data = st.session_state.logged_in_users[username]

    # Set current user context
    st.session_state.username = username
    st.session_state.is_admin = user_data['is_admin']
    st.session_state.file_paths = user_data['file_paths']

    # Load user's saved data if not already loaded
    if not user_data['apis']:
        try:
            # Always load APIs for any user (including admin)
            user_data['apis'] = load_user_apis(user_data['file_paths']["USER_APIS_FILE"])
            
            # Migration: Add module field to existing APIs that don't have it (default to EX)
            for api_name, api_config in user_data['apis'].items():
                if 'module' not in api_config:
                    api_config['module'] = 'EX'
            
            # Save the updated APIs with module information
            if user_data['apis']:
                save_user_apis(user_data['apis'], user_data['file_paths']["USER_APIS_FILE"])
        except Exception as e:
            st.error(f"Error loading user APIs for {username}: {str(e)}")
            user_data['apis'] = {}

    if not user_data['api_history']:
        user_data['api_history'] = load_api_history(user_data['file_paths']["API_HISTORY_FILE"])

    if not user_data['cookies_config']:
        if user_data['is_admin'] and username == "adminadmin":
            # For admin user, load from global admin cookies file
            if os.path.exists(ADMIN_COOKIES_FILE):
                try:
                    admin_cookies = load_api_configs(ADMIN_COOKIES_FILE)
                    # Initialize with admin cookies (not empty like regular users)
                    user_data['cookies_config'] = admin_cookies
                except Exception:
                    user_data['cookies_config'] = {}
            else:
                user_data['cookies_config'] = {}
        else:
            # For regular users, load with empty defaults
            user_data['cookies_config'] = load_cookies_config(
                user_data['file_paths']["COOKIES_CONFIG_FILE"],
                ADMIN_COOKIES_FILE
            )

    # Set session state to current user's data
    st.session_state.apis = user_data['apis']
    st.session_state.api_responses = user_data['api_responses']
    st.session_state.current_env = user_data['current_env']
    st.session_state.api_history = user_data['api_history']
    st.session_state.cookies_config = user_data['cookies_config']
    
    # Reset and recalculate module selection based on the new username
    if username.upper().strip() == "QA":
        st.session_state.selected_module = "AD"  # Only exact "QA" gets AD module
    else:
        st.session_state.selected_module = "EX"  # All others (including "QA xxx") get EX module
    
    # Always refresh cookies config to get latest admin values and ensure empty defaults
    if user_data['is_admin'] and username == "adminadmin":
        # For admin user, always load from global admin cookies file
        if os.path.exists(ADMIN_COOKIES_FILE):
            try:
                admin_cookies = load_api_configs(ADMIN_COOKIES_FILE)
                st.session_state.cookies_config = admin_cookies
                user_data['cookies_config'] = admin_cookies
            except Exception:
                st.session_state.cookies_config = {}
                user_data['cookies_config'] = {}
    else:
        # For regular users, load with empty defaults
        st.session_state.cookies_config = load_cookies_config(
            user_data['file_paths']["COOKIES_CONFIG_FILE"],
            ADMIN_COOKIES_FILE
        )
        # Update user data with refreshed cookies
        user_data['cookies_config'] = st.session_state.cookies_config


def _logout_user(username: str):
    """Logout a specific user"""
    if username in st.session_state.logged_in_users:
        # Save user's data before removing
        user_data = st.session_state.logged_in_users[username]
        if user_data['apis']:
            save_user_apis(user_data['apis'], user_data['file_paths']["USER_APIS_FILE"])

        # Remove user from logged in users
        del st.session_state.logged_in_users[username]

        # If this was the active user, switch to another user or clear session
        if st.session_state.active_user == username:
            if st.session_state.logged_in_users:
                # Prioritize QA users, then BA users when selecting a new active user
                qa_users = [user for user in st.session_state.logged_in_users.keys() 
                           if user.upper().startswith("QA")]
                ba_users = [user for user in st.session_state.logged_in_users.keys() 
                           if user.upper().startswith("BA")]
                regular_users = [user for user in st.session_state.logged_in_users.keys() 
                                if not user.upper().startswith("QA") and not user.upper().startswith("BA")]
                
                # Sort each group and combine with QA users first, then BA users
                qa_users.sort()
                ba_users.sort()
                regular_users.sort()
                sorted_users = qa_users + ba_users + regular_users
                
                # Switch to first available user (QA user if available, then BA user)
                new_active_user = sorted_users[0]
                st.session_state.active_user = new_active_user
                _load_user_data(new_active_user)
            else:
                # No users left, clear session
                st.session_state.active_user = None
                st.session_state.show_main_app = False
                for key in ['username', 'is_admin', 'file_paths', 'apis', 'api_responses', 
                           'current_env', 'api_history', 'cookies_config']:
                    if key in st.session_state:
                        del st.session_state[key]


def _save_current_user_data():
    """Save current user's data back to the logged_in_users dict"""
    if st.session_state.active_user and st.session_state.active_user in st.session_state.logged_in_users:
        user_data = st.session_state.logged_in_users[st.session_state.active_user]
        user_data['apis'] = st.session_state.get('apis', {})
        user_data['api_responses'] = st.session_state.get('api_responses', {})
        user_data['current_env'] = st.session_state.get('current_env', 'SIT')
        user_data['api_history'] = st.session_state.get('api_history', [])
        user_data['cookies_config'] = st.session_state.get('cookies_config', {})

def show_admin_panel():
    """Display admin panel for global environment and cookie configuration"""
    if not st.session_state.get("is_admin", False):
        st.warning("Admin access required")
        return

    st.title("Admin Panel - Environment & Cookie Management")

    # Load environments configuration
    environments = load_environments_config()
    
    # Create tabs for different admin functions
    tab_names = ["🌐 Environment Management", "🍪 Cookie Configuration", "🔧 API Configuration", "📝 Content Management", "⏰ Timer Run", "📁 File Management"]
    
    # Determine default tab index based on focus
    default_tab = 0
    if st.session_state.get("admin_tab_focus") == "timer_run":
        default_tab = 4  # Timer Run tab index
        # Clear the focus after setting
        if "admin_tab_focus" in st.session_state:
            del st.session_state["admin_tab_focus"]
    
    env_tab, cookie_tab, api_tab, content_tab, timer_tab, file_tab = st.tabs(tab_names)
    
    with env_tab:
        st.subheader("Environment Management")
        st.info("Manage environments that will be available to all users.")
        
        # Display existing environments
        st.write("**Current Environments:**")
        for env_name, config in environments.items():
            col1, col2, col3, col4 = st.columns([2, 4, 2, 1])
            
            with col1:
                status = "🟢 Enabled" if config.get('enabled', True) else "🔴 Disabled"
                st.write(f"**{env_name}**")
                st.write(status)
            
            with col2:
                st.text(config.get('base_url', ''))
            
            with col3:
                # Toggle enable/disable
                current_status = config.get('enabled', True)
                if st.button(f"{'Disable' if current_status else 'Enable'}", key=f"toggle_{env_name}"):
                    environments[env_name]['enabled'] = not current_status
                    if save_environments_config(environments):
                        st.success(f"Environment {env_name} {'disabled' if current_status else 'enabled'}")
                        st.rerun()
                    else:
                        st.error("Failed to update environment")
            
            with col4:
                # Delete environment (except core ones)
                if env_name not in ['SIT', 'DAI', 'UAT']:
                    if st.button("🗑️", key=f"delete_{env_name}", help="Delete environment"):
                        del environments[env_name]
                        if save_environments_config(environments):
                            st.success(f"Environment {env_name} deleted")
                            st.rerun()
                        else:
                            st.error("Failed to delete environment")
        
        st.markdown("---")
        
        # Add new environment
        st.subheader("Add New Environment")
        with st.form("add_environment_form"):
            new_env_name = st.text_input("Environment Name", help="e.g., PROD, DEV, STAGING")
            new_env_url = st.text_input("Base URL", help="e.g., https://api.example.com/v1")
            new_env_cookies = st.text_area("Default Cookies", help="Optional default cookies for this environment")
            
            submitted = st.form_submit_button("Add Environment")
            
            if submitted:
                if not new_env_name or not new_env_url:
                    st.error("Environment name and base URL are required")
                elif new_env_name.upper() in environments:
                    st.error(f"Environment {new_env_name.upper()} already exists")
                else:
                    # Add new environment
                    environments[new_env_name.upper()] = {
                        "name": new_env_name.upper(),
                        "base_url": new_env_url,
                        "default_cookies": new_env_cookies,
                        "enabled": True
                    }
                    
                    if save_environments_config(environments):
                        st.success(f"Environment {new_env_name.upper()} added successfully!")
                        st.rerun()
                    else:
                        st.error("Failed to add environment")
    
    with cookie_tab:
        st.subheader("Global Cookie Configuration")
        st.info("Configure global cookies that will be used by all users unless they override them.")
        
        # Load admin cookies
        admin_cookies = {}
        if os.path.exists(ADMIN_COOKIES_FILE):
            try:
                admin_cookies = load_api_configs(ADMIN_COOKIES_FILE)
            except Exception:
                admin_cookies = {}

        # Get enabled environments for cookie configuration
        enabled_envs = get_enabled_environments()
        
        if enabled_envs:
            # Create tabs for each enabled environment
            env_tabs = st.tabs(enabled_envs)

            for i, env in enumerate(enabled_envs):
                with env_tabs[i]:
                    st.subheader(f"{env} Environment Cookies")
                    
                    # Get current cookies (from admin file or environment default)
                    env_cookies_string = admin_cookies.get(env, "")
                    if not env_cookies_string and env in environments:
                        env_cookies_string = environments[env].get('default_cookies', '')
                    
                    # Display as editable text area
                    cookies_string = st.text_area(
                        f"Edit cookies for {env} (format: name1=value1; name2=value2)",
                        value=env_cookies_string,
                        height=150,
                        key=f"admin_cookies_{env}"
                    )
                    
                    # Save button for this environment
                    if st.button(f"Save {env} Cookies", key=f"save_admin_{env}"):
                        admin_cookies[env] = cookies_string
                        if save_cookies_config(admin_cookies, ADMIN_COOKIES_FILE):
                            st.success(f"Global cookies for {env} saved successfully!")
                            
                            # Clear all users' cookies for this environment so they use admin cookies by default
                            for username in st.session_state.logged_in_users:
                                user_data = st.session_state.logged_in_users[username]
                                # Clear this environment's cookies for the user
                                if user_data['cookies_config'].get(env):
                                    user_data['cookies_config'][env] = ""
                                
                                # Save user's updated cookies config to file
                                user_cookies_file = user_data['file_paths']["COOKIES_CONFIG_FILE"]
                                save_cookies_config(user_data['cookies_config'], user_cookies_file)
                            
                            # Also update current session if we're looking at this environment
                            if st.session_state.get('current_env') == env:
                                st.session_state.cookies_config[env] = ""
                            
                            st.info(f"All users' cookies for {env} have been cleared to use the new admin cookies.")
                        else:
                            st.error(f"Failed to save cookies for {env}")

            # Reset to defaults button
            if st.button("Reset All to Environment Defaults"):
                admin_cookies = {}
                for env in enabled_envs:
                    if env in environments:
                        admin_cookies[env] = environments[env].get('default_cookies', '')
                
                if save_cookies_config(admin_cookies, ADMIN_COOKIES_FILE):
                    st.success("All environments reset to default cookies")
                    
                    # Clear all users' cookies for all environments so they use the new admin defaults
                    for username in st.session_state.logged_in_users:
                        user_data = st.session_state.logged_in_users[username]
                        # Clear all environment cookies for the user
                        for env in enabled_envs:
                            user_data['cookies_config'][env] = ""
                        
                        # Save user's updated cookies config to file
                        user_cookies_file = user_data['file_paths']["COOKIES_CONFIG_FILE"]
                        save_cookies_config(user_data['cookies_config'], user_cookies_file)
                    
                    # Also update current session
                    for env in enabled_envs:
                        st.session_state.cookies_config[env] = ""
                    
                    st.info("All users' cookies have been cleared to use the new admin default cookies.")
                else:
                    st.error("Failed to reset cookies")
                st.rerun()
        else:
            st.warning("No enabled environments found. Please enable at least one environment in the Environment Management tab.")

    with api_tab:
        st.subheader("Predefined API Configuration")
        st.info("Manage predefined API configurations that are available to all users.")
        
        # Load current predefined APIs
        predefined_apis = load_api_configs(API_CONFIGS_FILE)
        
        # Two columns: List and Form
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("📋 Current Predefined APIs")
            
            if predefined_apis:
                # Filter and display
                module_filter = st.selectbox("Filter by Module", ["All", "EX", "AD"], key="api_filter")
                
                # Convert to ordered list for manipulation
                api_list = list(predefined_apis.items())
                
                # Filter APIs based on module
                if module_filter != "All":
                    api_list = [(name, config) for name, config in api_list 
                               if config.get("module", "").upper() == module_filter]
                
                if api_list:
                    st.info("💡 Use ⬆️ ⬇️ buttons to reorder APIs. Order affects how they appear in the predefined list.")
                    
                    for i, (api_name, api_config) in enumerate(api_list):
                        with st.expander(f"🔧 {api_name} ({api_config.get('module', 'Unknown')} Module)", expanded=False):
                            col_info, col_order, col_actions = st.columns([2.5, 0.8, 0.7])
                            
                            with col_info:
                                st.write(f"**Path:** `{api_config.get('path', api_config.get('url_path', 'N/A'))}`")
                                st.write(f"**Method:** `{api_config.get('method', 'N/A')}`")
                                st.write(f"**Module:** `{api_config.get('module', 'N/A')}`")
                                
                                if api_config.get('headers'):
                                    st.write("**Headers:**")
                                    st.code(json.dumps(api_config['headers'], indent=2), language='json')
                                
                                if api_config.get('body'):
                                    st.write("**Body Template:**")
                                    st.code(json.dumps(api_config['body'], indent=2), language='json')
                            
                            with col_order:
                                st.write("**Order:**")
                                # Move up button (disabled if first item)
                                if st.button("⬆️", key=f"up_{api_name}", disabled=(i == 0), help="Move up"):
                                    if i > 0:
                                        # Swap with previous item in the full predefined_apis dict
                                        all_api_list = list(predefined_apis.items())
                                        current_idx = next(j for j, (name, _) in enumerate(all_api_list) if name == api_name)
                                        if current_idx > 0:
                                            # Swap positions
                                            all_api_list[current_idx], all_api_list[current_idx - 1] = all_api_list[current_idx - 1], all_api_list[current_idx]
                                            # Rebuild ordered dict
                                            reordered_apis = {name: config for name, config in all_api_list}
                                            if save_api_config(reordered_apis, API_CONFIGS_FILE):
                                                st.success("Order updated!")
                                                st.rerun()
                                            else:
                                                st.error("Failed to update order")
                                
                                # Move down button (disabled if last item)
                                if st.button("⬇️", key=f"down_{api_name}", disabled=(i == len(api_list) - 1), help="Move down"):
                                    if i < len(api_list) - 1:
                                        # Swap with next item in the full predefined_apis dict
                                        all_api_list = list(predefined_apis.items())
                                        current_idx = next(j for j, (name, _) in enumerate(all_api_list) if name == api_name)
                                        if current_idx < len(all_api_list) - 1:
                                            # Swap positions
                                            all_api_list[current_idx], all_api_list[current_idx + 1] = all_api_list[current_idx + 1], all_api_list[current_idx]
                                            # Rebuild ordered dict
                                            reordered_apis = {name: config for name, config in all_api_list}
                                            if save_api_config(reordered_apis, API_CONFIGS_FILE):
                                                st.success("Order updated!")
                                                st.rerun()
                                            else:
                                                st.error("Failed to update order")
                            
                            with col_actions:
                                st.write("**Actions:**")
                                if st.button("Edit", key=f"edit_{api_name}"):
                                    st.session_state['edit_api_name'] = api_name
                                    st.session_state['edit_api_config'] = api_config.copy()
                                    st.rerun()
                                
                                if st.button("Delete", key=f"delete_{api_name}", type="secondary"):
                                    if st.session_state.get(f'confirm_delete_{api_name}', False):
                                        # Actually delete
                                        del predefined_apis[api_name]
                                        if save_api_config(predefined_apis, API_CONFIGS_FILE):
                                            st.success(f"API '{api_name}' deleted successfully!")
                                            if f'confirm_delete_{api_name}' in st.session_state:
                                                del st.session_state[f'confirm_delete_{api_name}']
                                            st.rerun()
                                        else:
                                            st.error("Failed to delete API")
                                    else:
                                        st.session_state[f'confirm_delete_{api_name}'] = True
                                        st.warning(f"Click delete again to confirm removal of '{api_name}'")
                                        st.rerun()
                else:
                    st.info(f"No APIs found for module: {module_filter}")
            else:
                st.warning("No predefined APIs found.")
        
        with col2:
            st.subheader("➕ Add/Edit API Configuration")
            
            # Add bulk reordering section first
            if predefined_apis:
                with st.expander("🔄 Bulk Reorder APIs", expanded=False):
                    st.info("Drag and drop to reorder all APIs at once")
                    
                    # Create a list of all APIs with their current order
                    all_apis = list(predefined_apis.keys())
                    
                    # Create a form for bulk reordering
                    with st.form("bulk_reorder_form"):
                        st.write("**Current Order (top to bottom):**")
                        
                        # Display current order as numbered list
                        for i, api_name in enumerate(all_apis):
                            module = predefined_apis[api_name].get('module', 'Unknown')
                            st.write(f"{i+1}. {api_name} ({module})")
                        
                        st.markdown("---")
                        
                        # Text area for new order
                        new_order_text = st.text_area(
                            "Enter new order (one API name per line):",
                            value="\n".join(all_apis),
                            height=200,
                            help="Copy the API names from above and rearrange them in your desired order. Make sure all names are spelled exactly as shown."
                        )
                        
                        # Submit button
                        if st.form_submit_button("Apply New Order"):
                            # Parse the new order
                            new_order = [line.strip() for line in new_order_text.split('\n') if line.strip()]
                            
                            # Validate that all APIs are present and no extras
                            if set(new_order) == set(all_apis) and len(new_order) == len(all_apis):
                                # Create new ordered dict
                                reordered_apis = {api_name: predefined_apis[api_name] for api_name in new_order}
                                
                                # Save the new order
                                if save_api_config(reordered_apis, API_CONFIGS_FILE):
                                    st.success("✅ API order updated successfully!")
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to save new API order")
                            else:
                                st.error("❌ Invalid order. Make sure all API names are included exactly once.")
                                # Show what's missing or extra
                                missing = set(all_apis) - set(new_order)
                                extra = set(new_order) - set(all_apis)
                                if missing:
                                    st.error(f"Missing APIs: {', '.join(missing)}")
                                if extra:
                                    st.error(f"Unknown APIs: {', '.join(extra)}")
                
                st.markdown("---")
            
            # Check if we're editing
            editing = st.session_state.get('edit_api_name', None)
            edit_config = st.session_state.get('edit_api_config', {})
            
            if editing:
                st.info(f"✏️ Editing: {editing}")
                if st.button("Cancel Edit"):
                    if 'edit_api_name' in st.session_state:
                        del st.session_state['edit_api_name']
                    if 'edit_api_config' in st.session_state:
                        del st.session_state['edit_api_config']
                    st.rerun()
            
            with st.form("api_config_form"):
                # API Name
                api_name = st.text_input(
                    "API Display Name",
                    value=editing if editing else "",
                    help="Friendly name for the API that users will see",
                    disabled=bool(editing)  # Don't allow changing name when editing
                )
                
                # API Path
                api_path = st.text_input(
                    "API Path",
                    value=edit_config.get('path', ''),
                    help="API endpoint path (e.g., /users)",
                    placeholder="/example"
                )
                
                # HTTP Method
                api_method = st.selectbox(
                    "HTTP Method",
                    ["GET", "POST", "PUT", "DELETE", "PATCH"],
                    index=["GET", "POST", "PUT", "DELETE", "PATCH"].index(edit_config.get('method', 'POST'))
                )
                
                # Module
                api_module = st.selectbox(
                    "Module",
                    ["EX", "AD"],
                    index=["EX", "AD"].index(edit_config.get('module', 'EX'))
                )
                
                # Headers
                headers_json = st.text_area(
                    "Headers (JSON)",
                    value=json.dumps(edit_config.get('headers', {
                        "Content-Type": "application/json"
                    }), indent=2),
                    height=100,
                    help="Headers in JSON format"
                )
                
                # Body Template
                body_json = st.text_area(
                    "Request Body Template (JSON)",
                    value=json.dumps(edit_config.get('body', {}), indent=2),
                    height=200,
                    help="Request body template in JSON format"
                )
                
                # Submit button
                submit_label = "Update API" if editing else "Add API"
                if st.form_submit_button(submit_label, type="primary"):
                    # Validation
                    if not api_name or not api_path:
                        st.error("API Name and Path are required!")
                    else:
                        try:
                            # Parse JSON inputs
                            headers_dict = json.loads(headers_json) if headers_json.strip() else {}
                            body_dict = json.loads(body_json) if body_json.strip() else {}
                            
                            # Check if API name already exists (only for new APIs)
                            if not editing and api_name in predefined_apis:
                                st.error(f"API '{api_name}' already exists! Choose a different name.")
                            else:
                                # Create API config
                                new_config = {
                                    "path": api_path.strip(),
                                    "method": api_method,
                                    "module": api_module,
                                    "headers": headers_dict,
                                    "body": body_dict
                                }
                                
                                if editing:
                                    # Update existing API
                                    predefined_apis[editing] = new_config
                                    action = f"API '{editing}' updated"
                                else:
                                    # Add new API
                                    predefined_apis[api_name] = new_config
                                    action = f"API '{api_name}' added"
                                
                                # Save to file
                                if save_api_config(predefined_apis, API_CONFIGS_FILE):
                                    st.success(f"{action} successfully!")
                                    
                                    # Clear edit state
                                    if 'edit_api_name' in st.session_state:
                                        del st.session_state['edit_api_name']
                                    if 'edit_api_config' in st.session_state:
                                        del st.session_state['edit_api_config']
                                    
                                    st.rerun()
                                else:
                                    st.error(f"Failed to save {action.lower()}")
                        
                        except json.JSONDecodeError as e:
                            st.error(f"Invalid JSON format: {str(e)}")
                        except Exception as e:
                            st.error(f"Error: {str(e)}")

    with content_tab:
        st.subheader("Content Management")
        st.info("Manage the admin-editable content that users can view from the login page.")
        
        # Load current admin content
        current_content = load_admin_content()
        
        # Display current content info
        content_lines = len(current_content.split('\n'))
        content_chars = len(current_content)
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Lines", content_lines)
        with col2:
            st.metric("Characters", content_chars)
        
        st.markdown("---")
        
        # Content editor
        st.subheader("✏️ Edit Admin Content")
        
        edited_content = st.text_area(
            "Content (Markdown Format)",
            value=current_content,
            height=300,
            help="This content will be visible to all users when they click 'Manage Content' on the login page."
        )
        
        # Action buttons
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Save Content", type="primary"):
                if save_admin_content(edited_content):
                    st.success("Admin content saved successfully!")
                    st.rerun()
                else:
                    st.error("Failed to save admin content")
        
        with col2:
            if st.button("🔄 Reset to Default"):
                default_content = "# Admin Editable Content\n\nThis content can be edited by administrators.\n\n## Welcome\n\nWelcome to the API Client Tester!\n\n## Getting Started\n\nPlease contact your administrator for more information."
                if save_admin_content(default_content):
                    st.success("Content reset to default!")
                    st.rerun()
        
        with col3:
            # Preview toggle
            show_preview = st.checkbox("Show Preview", value=False)
        
        if show_preview:
            st.markdown("---")
            st.subheader("📄 Content Preview")
            with st.container():
                st.markdown(edited_content)

    with timer_tab:
        st.subheader("Timer Run Management")
        st.info("Configure Timer Job URLs for easy copy & paste execution by users.")
        
        # Manual Timer Job URL Generator Section
        with st.expander("📝 Timer Job URL Generator", expanded=True):
            st.write("**Create Timer Job URLs for SIT environment:**")
            st.info("URLs will be automatically saved to ADMIN_CONTENT.md for public viewing")
            
            # Initialize timer job entries in session state
            if 'timer_job_entries' not in st.session_state:
                st.session_state.timer_job_entries = []
            
            # Form to add new timer job entry
            with st.form("add_timer_job_form"):
                col1, col2 = st.columns(2)
                
                with col1:
                    job_name = st.text_input(
                        "Timer Job Name",
                        placeholder="e.g., Daily Backup Job",
                        help="Descriptive name for this timer job"
                    )
                
                with col2:
                    job_id = st.text_input(
                        "Timer Job ID",
                        placeholder="e.g., job123456",
                        help="The actual timer job ID to replace in URLs"
                    )
                
                # API Path configuration
                api_path = st.text_input(
                    "API Path Template",
                    value="/DEVTimerJob/DEVTriggerTimerJob/{timer_job_id}",
                    help="API path with {timer_job_id} placeholder"
                )
                
                # Module selection
                module = st.selectbox(
                    "Module",
                    ["EX", "AD"],
                    help="EX: Assessment Module, AD: Administration Module"
                )
                
                submitted = st.form_submit_button("➕ Add Timer Job Entry")
                
                if submitted and job_name and job_id:
                    # Add to session state
                    new_entry = {
                        'name': job_name,
                        'id': job_id,
                        'path': api_path,
                        'module': module
                    }
                    st.session_state.timer_job_entries.append(new_entry)
                    
                    # Auto-update ADMIN_CONTENT.md
                    _update_admin_content_with_timer_jobs()
                    
                    st.success(f"Added timer job entry: {job_name}")
                    st.success("✅ ADMIN_CONTENT.md updated automatically!")
                    st.rerun()
            
            # Display existing entries with SIT URLs only
            if st.session_state.timer_job_entries:
                st.markdown("---")
                st.write("**Generated Timer Job URLs (SIT Environment):**")
                
                # Check if SIT environment exists
                enabled_envs = get_enabled_environments()
                
                if "SIT" in enabled_envs:
                    for i, entry in enumerate(st.session_state.timer_job_entries):
                        with st.container():
                            # Entry header with edit/remove buttons
                            col_header, col_edit, col_remove, col_update = st.columns([4, 1, 1, 2])
                            
                            with col_header:
                                st.write(f"**🎯 {entry['name']}** (ID: `{entry['id']}`)")
                            
                            with col_edit:
                                if st.button("✏️", key=f"edit_timer_entry_{i}", help="Edit entry"):
                                    st.session_state[f"editing_timer_{i}"] = True
                                    st.rerun()
                            
                            with col_remove:
                                if st.button("🗑️", key=f"remove_timer_entry_{i}", help="Remove entry"):
                                    st.session_state.timer_job_entries.pop(i)
                                    # Auto-update ADMIN_CONTENT.md after removal
                                    _update_admin_content_with_timer_jobs()
                                    st.success("Timer job entry removed and ADMIN_CONTENT.md updated!")
                                    st.rerun()
                            
                            with col_update:
                                if st.button("🔄 Update Content", key=f"update_content_{i}", help="Update ADMIN_CONTENT.md"):
                                    _update_admin_content_with_timer_jobs()
                                    st.success("✅ ADMIN_CONTENT.md updated!")
                            
                            # Inline editing form
                            if st.session_state.get(f"editing_timer_{i}", False):
                                with st.form(f"edit_timer_form_{i}"):
                                    edit_name = st.text_input("Name", value=entry['name'], key=f"edit_name_{i}")
                                    edit_id = st.text_input("ID", value=entry['id'], key=f"edit_id_{i}")
                                    edit_path = st.text_input("Path", value=entry['path'], key=f"edit_path_{i}")
                                    edit_module = st.selectbox("Module", ["EX", "AD"], 
                                                             index=0 if entry['module'] == 'EX' else 1, 
                                                             key=f"edit_module_{i}")
                                    
                                    col_save, col_cancel = st.columns(2)
                                    with col_save:
                                        if st.form_submit_button("💾 Save & Update"):
                                            st.session_state.timer_job_entries[i] = {
                                                'name': edit_name,
                                                'id': edit_id,
                                                'path': edit_path,
                                                'module': edit_module
                                            }
                                            st.session_state[f"editing_timer_{i}"] = False
                                            
                                            # Auto-update ADMIN_CONTENT.md after edit
                                            _update_admin_content_with_timer_jobs()
                                            
                                            st.success("Timer job entry updated!")
                                            st.success("✅ ADMIN_CONTENT.md updated automatically!")
                                            st.rerun()
                                    
                                    with col_cancel:
                                        if st.form_submit_button("❌ Cancel"):
                                            st.session_state[f"editing_timer_{i}"] = False
                                            st.rerun()
                            else:
                                # Display SIT URL only
                                base_url = get_current_base_url("SIT", entry['module'])
                                final_path = entry['path'].replace('{timer_job_id}', entry['id'])
                                full_url = f"{base_url}{final_path}"
                                
                                # Display with SIT environment label and copy-friendly format
                                st.text_input(
                                    f"SIT Environment URL:",
                                    value=full_url,
                                    key=f"url_{entry['name']}_SIT_{i}",
                                    help=f"Copy this URL to run {entry['name']} in SIT"
                                )
                            
                            st.markdown("---")
                        
                    # Auto-update all content button
                    if st.button("🔄 Update All to ADMIN_CONTENT.md", type="primary"):
                        _update_admin_content_with_timer_jobs()
                        st.success("✅ All timer job entries updated to ADMIN_CONTENT.md!")
                        
                else:
                    st.warning("⚠️ SIT environment not found. Configure SIT environment first in the Environment Management tab.")
            else:
                st.info("💡 Add timer job entries above to generate SIT URLs and auto-update ADMIN_CONTENT.md.")
        
        # Predefined Timer Job APIs Section (from API configuration)
        with st.expander("⚙️ Predefined Timer Job APIs", expanded=False):
            st.write("**Configure Timer Job APIs for Public Display:**")
            
            # Load predefined Timer Job API if available
            predefined_apis = load_api_configs(API_CONFIGS_FILE)
            timer_job_apis = {name: config for name, config in predefined_apis.items() 
                            if "timer" in name.lower() or "DEVTriggerTimerJob" in config.get('path', '')}
            
            if timer_job_apis:
                st.success(f"Found {len(timer_job_apis)} Timer Job API(s) configured")
                
                # Add action buttons
                col1, col2, col3 = st.columns(3)
                with col1:
                    if st.button("📄 Generate Public Info", help="Generate markdown info for users"):
                        markdown_content = generate_timer_job_markdown()
                        st.success("Timer Job information generated!")
                        with st.expander("Generated Markdown (Preview)", expanded=True):
                            st.markdown(markdown_content)
                
                with col2:
                    if st.button("🔧 Configure New API", help="Add Timer Job API in API Configuration tab"):
                        st.info("💡 Switch to the 'API Configuration' tab to add/modify Timer Job APIs")
                
                with col3:
                    if st.button("👀 Preview Public View", help="See what users will see"):
                        show_timer_job_dialog()
                
                st.markdown("---")
                
                # List existing Timer Job APIs
                for api_name, config in timer_job_apis.items():
                    with st.container():
                        col1, col2 = st.columns([3, 1])
                        with col1:
                            st.write(f"**{api_name}**")
                            st.write(f"Path: `{config.get('path', 'N/A')}`")
                            st.write(f"Method: `{config.get('method', 'N/A')}`")
                            st.write(f"Module: `{config.get('module', 'EX')}`")
                        with col2:
                            if st.button(f"✏️ Edit", key=f"timer_edit_{api_name}"):
                                st.info(f"💡 Use the 'API Configuration' tab to modify {api_name}")
                        st.markdown("---")
            else:
                st.warning("No Timer Job APIs found in predefined configurations")
                st.info("💡 **Tip**: Add a Timer Job API in the API Configuration tab first")
                
                if st.button("➕ Add Timer Job API", type="primary"):
                    st.info("💡 Switch to the 'API Configuration' tab to add Timer Job APIs")
        
        # Public Information Section
        with st.expander("📋 Public Timer Job Information", expanded=True):
            st.write("**What users see when they click 'Timer Run' button:**")
            
            # Generate and display current markdown
            current_markdown = generate_timer_job_markdown()
            
            col1, col2 = st.columns([2, 1])
            with col1:
                st.markdown(current_markdown)
            
            with col2:
                st.info("**Auto-Generated Content**\n\nThis content is automatically generated from your Timer Job API configurations and will be shown to all users.")
                
                if st.button("🔄 Refresh Preview"):
                    st.rerun()
        
        # Environment URLs Generation Section
        with st.expander("🌐 Environment URLs", expanded=False):
            st.write("**Generate Timer Job URLs for all environments:**")
            
            enabled_envs = get_enabled_environments()
            if enabled_envs and timer_job_apis:
                for api_name, config in timer_job_apis.items():
                    st.write(f"**{api_name}:**")
                    
                    for env in enabled_envs:
                        base_url = get_current_base_url(env, config.get('module', 'EX'))
                        path = config.get('path', '')
                        
                        if '{timer_job_id}' in path:
                            sample_url = f"{base_url}{path}".replace('{timer_job_id}', '[TIMER_JOB_ID]')
                        else:
                            sample_url = f"{base_url}{path}"
                        
                        st.code(f"{env}: {sample_url}", language="text")
                    
                    st.markdown("---")
            else:
                st.warning("Configure Timer Job APIs and environments first.")
        
        # Configuration Management Section
        with st.expander("⚙️ Configuration Management", expanded=False):
            st.write("**Timer Job API Management Tools:**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Quick Actions:**")
                if st.button("📤 Export Timer Job Config", help="Export Timer Job APIs configuration"):
                    if timer_job_apis:
                        st.json(timer_job_apis)
                        st.success("Timer Job configuration displayed above")
                    else:
                        st.warning("No Timer Job APIs to export")
                
                if st.button("🔄 Refresh Configuration", help="Reload Timer Job APIs from file"):
                    st.success("Configuration refreshed!")
                    st.rerun()
            
            with col2:
                st.write("**Information:**")
                st.info(f"**Total APIs:** {len(timer_job_apis)}")
                st.info(f"**Environments:** {len(get_enabled_environments())}")
                
                if timer_job_apis:
                    modules_used = set(config.get('module', 'EX') for config in timer_job_apis.values())
                    st.info(f"**Modules:** {', '.join(modules_used)}")
            
            st.markdown("---")
            st.write("**💡 Tips:**")
            st.write("- Timer Job APIs configured here automatically appear in the public Timer Run dialog")
            st.write("- Users can view Timer Job information by clicking the Timer Run button")
            st.write("- Environment URLs are automatically generated for all configured environments")
            st.write("- Use the API Configuration tab to add, edit, or remove Timer Job APIs")

    with file_tab:
        st.subheader("File Management")
        st.info("Manage uploaded Excel files from all users.")
        
        upload_dir = os.path.join(os.path.dirname(__file__), "upload_data")
        if os.path.exists(upload_dir):
            all_files = [f for f in os.listdir(upload_dir) if f.endswith(('.xlsx', '.xls'))]
            
            if all_files:
                st.success(f"📊 Found {len(all_files)} uploaded file(s) in total")
                
                # Summary stats
                total_size = 0
                for file_name in all_files:
                    try:
                        file_path = os.path.join(upload_dir, file_name)
                        total_size += os.path.getsize(file_path)
                    except:
                        pass
                
                total_size_mb = round(total_size / (1024 * 1024), 2)
                st.info(f"💾 Total storage used: {total_size_mb} MB")
                
                # Group files by user
                user_files = {}
                for file_name in all_files:
                    # Extract username from filename (username_apiname_timestamp.extension)
                    parts = file_name.split('_')
                    if len(parts) >= 2:
                        username_part = parts[0]
                        if username_part not in user_files:
                            user_files[username_part] = []
                        user_files[username_part].append(file_name)
                    else:
                        # Files that don't follow naming convention
                        if 'other' not in user_files:
                            user_files['other'] = []
                        user_files['other'].append(file_name)
                
                # Sort users - prioritize problematic ones
                sorted_users = sorted(user_files.keys())
                
                for user in sorted_users:
                    files = user_files[user]
                    user_size = 0
                    for file_name in files:
                        try:
                            file_path = os.path.join(upload_dir, file_name)
                            user_size += os.path.getsize(file_path)
                        except:
                            pass
                    user_size_mb = round(user_size / (1024 * 1024), 2)
                    
                    with st.expander(f"👤 **{user}** ({len(files)} files, {user_size_mb} MB)", expanded=False):
                        
                        for file_name in sorted(files, reverse=True):  # Most recent first
                            file_path = os.path.join(upload_dir, file_name)
                            try:
                                file_size = os.path.getsize(file_path)
                                file_size_kb = round(file_size / 1024, 2)
                                
                                # Extract timestamp from filename
                                try:
                                    name_parts = file_name.split('_')
                                    if len(name_parts) >= 3:
                                        date_part = name_parts[-2]
                                        time_part = name_parts[-1].split('.')[0]
                                        timestamp_str = f"{date_part}_{time_part}"
                                        timestamp = datetime.datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                                        formatted_time = timestamp.strftime("%Y-%m-%d %H:%M:%S")
                                    else:
                                        formatted_time = "Unknown time"
                                except:
                                    formatted_time = "Unknown time"
                                
                                col1, col2 = st.columns([4, 1])
                                with col1:
                                    st.text(f"📄 {file_name}")
                                    st.caption(f"Uploaded: {formatted_time} | Size: {file_size_kb} KB")
                                
                                with col2:
                                    if st.button("🗑️", key=f"admin_delete_{file_name}", help="Delete file"):
                                        try:
                                            os.remove(file_path)
                                            st.success(f"✅ Deleted {file_name}")
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"❌ Error deleting file: {str(e)}")
                            except Exception as e:
                                st.error(f"Error accessing file {file_name}: {str(e)}")
                        
                        # Delete all files for this user
                        if st.button(f"🗑️ Delete all files for {user}", key=f"delete_user_files_{user}"):
                            deleted_count = 0
                            failed_count = 0
                            
                            for file_name in files:
                                try:
                                    file_path = os.path.join(upload_dir, file_name)
                                    os.remove(file_path)
                                    deleted_count += 1
                                except Exception:
                                    failed_count += 1
                            
                            if deleted_count > 0:
                                st.success(f"✅ Deleted {deleted_count} file(s) for {user}")
                            if failed_count > 0:
                                st.error(f"❌ Failed to delete {failed_count} file(s)")
                            
                            if deleted_count > 0:
                                st.rerun()
                
                st.markdown("---")
                
                # Bulk operations
                st.subheader("🗑️ Bulk Operations")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Delete all files
                    if st.button("🗑️ Delete ALL Files", key="admin_delete_all_files", type="secondary"):
                        deleted_count = 0
                        failed_count = 0
                        
                        for file_name in all_files:
                            try:
                                file_path = os.path.join(upload_dir, file_name)
                                os.remove(file_path)
                                deleted_count += 1
                            except Exception:
                                failed_count += 1
                        
                        if deleted_count > 0:
                            st.success(f"✅ Deleted {deleted_count} file(s)")
                        if failed_count > 0:
                            st.error(f"❌ Failed to delete {failed_count} file(s)")
                        
                        if deleted_count > 0:
                            st.rerun()
                
                with col2:
                    # Delete files older than X days
                    days_old = st.number_input("Delete files older than (days):", min_value=1, max_value=365, value=30)
                    if st.button(f"🗑️ Delete files older than {days_old} days", key="delete_old_files"):
                        import datetime as dt
                        cutoff_date = dt.datetime.now() - dt.timedelta(days=days_old)
                        
                        deleted_count = 0
                        failed_count = 0
                        
                        for file_name in all_files:
                            try:
                                file_path = os.path.join(upload_dir, file_name)
                                file_mtime = dt.datetime.fromtimestamp(os.path.getmtime(file_path))
                                
                                if file_mtime < cutoff_date:
                                    os.remove(file_path)
                                    deleted_count += 1
                            except Exception:
                                failed_count += 1
                        
                        if deleted_count > 0:
                            st.success(f"✅ Deleted {deleted_count} old file(s)")
                        elif failed_count > 0:
                            st.error(f"❌ Failed to delete {failed_count} file(s)")
                        else:
                            st.info("No files older than the specified period found")
                        
                        if deleted_count > 0:
                            st.rerun()
            else:
                st.info("📂 No uploaded files found")
                st.write("Files will appear here when users upload Excel files through the AD module APIs.")
        else:
            st.warning("📁 Upload directory does not exist")
            st.write("The upload directory will be created automatically when the first file is uploaded.")

    if st.button("Back to API Tester"):
        st.session_state.admin_mode = False
        st.rerun()
def main():
    """Main."""
    # Initialize session state
    if "logged_in_users" not in st.session_state:
        st.session_state.logged_in_users = {}
    if "active_user" not in st.session_state:
        st.session_state.active_user = None
    if "show_main_app" not in st.session_state:
        st.session_state.show_main_app = False

    # Show login page if no users are logged in or main app not started
    if not st.session_state.logged_in_users or not st.session_state.show_main_app:
        show_multi_user_login()
        return

    # Check if admin mode is active
    if (st.session_state.get("is_admin", False) and 
        st.session_state.get("admin_mode", False)):
        show_admin_panel()
        return

    # Save current user data before any operations
    _save_current_user_data()

    # Get file paths for current user
    file_paths = st.session_state.file_paths

    # Help and content buttons above title
    col1, col2, col3 = st.columns([1, 1, 8])
    
    with col1:
        if st.button("📖 Instruction", help="Take a look about instruction"):
            show_help_dialog()
    
    with col2:
        if st.button("⏰ Timer Run", help="View Timer Job APIs information"):
            show_timer_job_dialog()

    # Create a layout with title and user switcher
    title_col, user_col, _ = st.columns([2, 3, 3])

    with title_col:
        # Determine default module based on username if not set
        username = st.session_state.get('username', '')
        if username.upper() == "QA":
            default_module = "AD"
        else:
            default_module = "EX"
        
        selected_module = st.session_state.get('selected_module', default_module)
        module_name = "Assessment" if selected_module == "EX" else "Administration"
        st.markdown(
            f"<h3 style='margin:0;'>API Tester - {st.session_state.get('username', 'Unknown')}</h3>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<p style='margin:0; color: #666;'>Module: {module_name} ({selected_module})</p>",
            unsafe_allow_html=True
        )

    with user_col:
        # Module selection dropdown
        module_options = ["EX", "AD"]
        
        # Initialize module in session state if not exists
        if 'selected_module' not in st.session_state:
            # Determine default module based on username
            username = st.session_state.get('username', '')
            if username.upper().strip() == "QA":
                st.session_state.selected_module = "AD"  # QA accounts default to AD (Administration)
            else:
                st.session_state.selected_module = "EX"  # Other accounts default to EX (Assessment)
        
        current_module_index = module_options.index(st.session_state.selected_module) if st.session_state.selected_module in module_options else 0

        selected_module = st.selectbox(
            "Module",
            module_options,
            index=current_module_index,
            key="module_switcher",
            help="EX: Assessment Module, AD: Administration Module"
        )

        if selected_module != st.session_state.selected_module:
            st.session_state.selected_module = selected_module
            
            # Update URLs only for APIs that match the newly selected module
            for api_name in st.session_state.apis:
                api_config = st.session_state.apis[api_name]
                api_module = api_config.get('module', 'EX')
                if api_module == selected_module and "path" in api_config:
                    path = api_config["path"]
                    api_config["url"] = f"{get_current_base_url(st.session_state.current_env, selected_module)}{path}"

            st.rerun()

    # Initialize session state for current user
    if 'api_responses' not in st.session_state:
        st.session_state.api_responses = {}
    if 'current_env' not in st.session_state:
        # Set default environment based on username (BA users default to DEMO, QA users default to SIT)
        enabled_envs = get_enabled_environments()
        username = st.session_state.get('username', '')
        if username.upper().startswith("BA") and "DEMO" in enabled_envs:
            st.session_state.current_env = "DEMO"
        else:
            st.session_state.current_env = "SIT" if "SIT" in enabled_envs else enabled_envs[0] if enabled_envs else "SIT"

    # Admin options and logout in sidebar
    username = st.session_state.get('username', 'Unknown')
    st.sidebar.markdown(f"**Active User:** {username}")
    
    # User switcher in sidebar
    if len(st.session_state.logged_in_users) > 1:
        st.sidebar.subheader("Switch User")
        user_options = list(st.session_state.logged_in_users.keys())
        
        # Sort the user options to prioritize QA users, then BA users
        qa_users = [user for user in user_options if user.upper().startswith("QA")]
        ba_users = [user for user in user_options if user.upper().startswith("BA")]
        regular_users = [user for user in user_options if not user.upper().startswith("QA") and not user.upper().startswith("BA")]
        
        # Sort each group alphabetically and combine with QA users first, then BA users
        qa_users.sort()
        ba_users.sort()
        regular_users.sort()
        sorted_user_options = qa_users + ba_users + regular_users
        
        # Find the current user's index in the sorted list
        current_index = sorted_user_options.index(st.session_state.active_user) if st.session_state.active_user in sorted_user_options else 0

        selected_user = st.sidebar.selectbox(
            "Switch to:",
            sorted_user_options,
            index=current_index,
            key="sidebar_user_switcher"
        )

        if selected_user != st.session_state.active_user:
            st.session_state.active_user = selected_user
            _load_user_data(selected_user)
            st.rerun()

    if st.session_state.get("is_admin", False):
        if st.sidebar.button("Admin Panel"):
            st.session_state.admin_mode = True
            st.rerun()

    # Individual user logout
    username = st.session_state.get('username', 'Unknown')
    if st.sidebar.button(f"🚪 Logout {username}"):
        _logout_user(st.session_state.username)
        st.rerun()

    # Environment selector - use enabled environments
    enabled_envs = get_enabled_environments()
    current_env_index = 0
    
    if st.session_state.current_env in enabled_envs:
        current_env_index = enabled_envs.index(st.session_state.current_env)
    elif enabled_envs:
        # If current env is not available, default to first enabled env
        st.session_state.current_env = enabled_envs[0]
        current_env_index = 0
    
    if enabled_envs:
        env = st.sidebar.selectbox(
            "Environment", 
            enabled_envs,
            index=current_env_index
        )
        
        if env != st.session_state.current_env:
            st.session_state.current_env = env
            # Update current user's environment
            if st.session_state.active_user in st.session_state.logged_in_users:
                st.session_state.logged_in_users[st.session_state.active_user]['current_env'] = env

            # Refresh cookies config to get latest admin values for all environments
            st.session_state.cookies_config = load_cookies_config(
                st.session_state.file_paths["COOKIES_CONFIG_FILE"],
                ADMIN_COOKIES_FILE
            )

            # Update URLs for all APIs to use the new environment's base URL
            for api_name in st.session_state.apis:
                if "path" in st.session_state.apis[api_name]:
                    path = st.session_state.apis[api_name]["path"]
                    api_module = st.session_state.apis[api_name].get('module', 'EX')  # Use saved module or default to EX
                    st.session_state.apis[api_name]["url"] = f"{get_current_base_url(env, api_module)}{path}"

            st.rerun()
    else:
        st.sidebar.warning("No environments enabled. Contact admin.")

    # Show current base URL for selected module
    # Determine default module based on username if not set
    username = st.session_state.get('username', '')
    if username.upper() == "QA":
        default_module = "AD"
    else:
        default_module = "EX"
    
    selected_module = st.session_state.get('selected_module', default_module)
    module_name = "Assessment" if selected_module == "EX" else "Administration"
    st.sidebar.info(f"Base URL ({module_name}): {get_current_base_url(st.session_state.current_env, selected_module)}")

    # Cookies management section in sidebar
    with st.sidebar.expander("Manage Environment Cookies"):
        manage_cookies(file_paths["COOKIES_CONFIG_FILE"])

    # Create a two-column layout
    col1, col2 = st.columns([1, 3])

    with col1:
        # Determine if user is a QA or BA account (username starts with QA or BA)
        username = st.session_state.get('username', '')
        is_priority_account = username.upper().startswith("QA") or username.upper().startswith("BA")
        
        # Combined API Management and Predefined API Tests in a single expander
        # Auto-close for QA/BA accounts, open for others
        with st.expander("API Management & Predefined Tests", expanded=not is_priority_account):
            # API Management section
            st.subheader("API Management")
            add_new_api()
            
            # Predefined API Tests section
            st.subheader("Predefined API Tests")
            load_predefined_api(file_paths["API_CONFIG_FILE"])

        # List of saved APIs
        if st.session_state.apis:
            # Filter APIs by selected module
            # Determine default module based on username if not set
            username = st.session_state.get('username', '')
            if username.upper() == "QA":
                default_module = "AD"
            else:
                default_module = "EX"
            
            selected_module = st.session_state.get('selected_module', default_module)
            filtered_apis = {
                name: config for name, config in st.session_state.apis.items() 
                if config.get('module', 'EX') == selected_module
            }
            
            if filtered_apis:
                # Saved APIs in an expander - always expanded regardless of account type
                with st.expander(f"Saved APIs ({selected_module})", expanded=True):
                    # Create a container with fixed height for scrolling
                    api_list_container = st.container()

                    # Set max height with CSS to enable scrolling
                    api_list_container.markdown("""
                    <style>
                    .api-list {
                        max-height: 300px;
                        overflow-y: auto;
                        padding-right: 10px;
                    }
                    </style>
                    """, unsafe_allow_html=True)

                    # Create scrollable area
                    with api_list_container:
                        st.markdown('<div class="api-list">', unsafe_allow_html=True)

                        # Display each API with open, rename and delete buttons
                        # Use reversed list to show newest APIs first
                        api_names = list(filtered_apis.keys())
                        # Reverse the order so newest APIs appear first
                        for api_name in reversed(api_names):
                            cols = st.columns([3, 1, 1, 1])

                            # Display API name
                            cols[0].write(api_name)

                            # Open button
                            if cols[1].button("🔍", key=f"open_{api_name}"):
                                # No need to clear API-specific keys anymore as we use unique keys per API
                                    
                                st.session_state.current_api = api_name
                                st.rerun()

                            # Rename button
                            if cols[2].button("✏️", key=f"rename_{api_name}"):
                                # Show rename form
                                st.session_state[f"rename_mode_{api_name}"] = True
                                st.rerun()

                            # Delete button
                            if cols[3].button("🗑️", key=f"del_{api_name}"):
                                # Delete the API
                                del st.session_state.apis[api_name]

                                # Also delete any associated response
                                if api_name in st.session_state.api_responses:
                                    del st.session_state.api_responses[api_name]

                                # Update current_api if needed
                                if 'current_api' in st.session_state and st.session_state.current_api == api_name:
                                    # Find another API of the same module to switch to
                                    remaining_apis = {
                                        name: config for name, config in st.session_state.apis.items() 
                                        if config.get('module', 'EX') == selected_module
                                    }
                                    if remaining_apis:
                                        st.session_state.current_api = next(iter(remaining_apis))
                                    else:
                                        if 'current_api' in st.session_state:
                                            del st.session_state.current_api

                                # Save the updated APIs to file
                                save_user_apis(st.session_state.apis, st.session_state.file_paths["USER_APIS_FILE"])

                                st.success(f"Deleted API: {api_name}")
                                st.rerun()

                            # Show rename form if in rename mode
                            if st.session_state.get(f"rename_mode_{api_name}", False):
                                with st.form(key=f"rename_form_{api_name}"):
                                    new_name = st.text_input("New name", value=api_name)

                                    rename_cols = st.columns(2)
                                    rename_submit = rename_cols[0].form_submit_button("Save")
                                    rename_cancel = rename_cols[1].form_submit_button("Cancel")

                                    if rename_submit and new_name:
                                        if new_name != api_name:
                                            if new_name in st.session_state.apis:
                                                st.error(f"Name '{new_name}' already exists")
                                            else:
                                                # Create a copy of the API with the new name
                                                st.session_state.apis[new_name] = st.session_state.apis[api_name].copy()

                                                # Copy response data if exists
                                                if api_name in st.session_state.api_responses:
                                                    st.session_state.api_responses[new_name] = st.session_state.api_responses[api_name]

                                                # Delete old API
                                                del st.session_state.apis[api_name]

                                                # Update current_api if needed
                                                if 'current_api' in st.session_state and st.session_state.current_api == api_name:
                                                    st.session_state.current_api = new_name

                                                # Save updated APIs to file
                                                save_user_apis(st.session_state.apis, st.session_state.file_paths["USER_APIS_FILE"])

                                                st.success(f"Renamed '{api_name}' to '{new_name}'")

                                                # Reset rename mode
                                                del st.session_state[f"rename_mode_{api_name}"]
                                                st.rerun()

                                    if rename_cancel:
                                        # Reset rename mode
                                        del st.session_state[f"rename_mode_{api_name}"]
                                        st.rerun()

                            # Add a separator line
                            st.markdown("---")

                        st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.info(f"No APIs saved for {selected_module} module yet. Add one above or load from predefined tests.")
        else:
            st.info("No APIs added yet. Add one above or load from predefined tests.")

    with col2:
        if 'current_api' in st.session_state:
            if st.session_state.current_api in st.session_state.apis:
                # Check if the current API matches the selected module
                current_api_config = st.session_state.apis[st.session_state.current_api]
                current_api_module = current_api_config.get('module', 'EX')
                
                # Determine default module based on username if not set
                username = st.session_state.get('username', '')
                if username.upper() == "QA":
                    default_module = "AD"
                else:
                    default_module = "EX"
                
                selected_module = st.session_state.get('selected_module', default_module)
                
                if current_api_module == selected_module:
                    display_api_tester(st.session_state.current_api, file_paths)
                else:
                    # API doesn't match current module, offer to switch
                    st.info(f"Current API is from {current_api_module} module. Switch to {current_api_module} module to view it.")
                    if st.button(f"Switch to {current_api_module} Module"):
                        st.session_state.selected_module = current_api_module
                        st.rerun()
            else:
                st.info("Select an API to test or add a new one.")
        else:
            # Determine default module based on username if not set
            username = st.session_state.get('username', '')
            if username.upper() == "QA":
                default_module = "AD"
            else:
                default_module = "EX"
            
            selected_module = st.session_state.get('selected_module', default_module)
            module_name = "Assessment" if selected_module == "EX" else "Administration"
            st.info(f"Select an API from the {module_name} ({selected_module}) module to test or add a new one.")


def manage_cookies(file_path):
    """Manage cookies for each environment using string format"""
    st.subheader("Cookies for " + st.session_state.current_env)
    
    # Check if current user is admin
    is_admin = st.session_state.get("is_admin", False)
    
    if is_admin:
        # Admin user - manage global cookies
        st.info("🔧 **Admin Mode**: You are managing global cookies that affect all users.")
        
        # Load admin cookies
        admin_cookies = {}
        if os.path.exists(ADMIN_COOKIES_FILE):
            try:
                admin_cookies = load_api_configs(ADMIN_COOKIES_FILE)
            except Exception:
                admin_cookies = {}

        # Get current environment cookies string from admin config
        env_cookies_string = admin_cookies.get(st.session_state.current_env, "")
        
        # Display as editable text area with admin context
        cookies_string = st.text_area(
            f"Global cookies for {st.session_state.current_env} (affects all users)",
            value=env_cookies_string,
            height=150,
            help="These cookies will be used by all users unless they set custom cookies"
        )

        # Save button for admin
        if st.button("Save Global Cookies"):
            admin_cookies[st.session_state.current_env] = cookies_string
            print(f"[DEBUG] Admin saving cookies for {st.session_state.current_env}: {cookies_string}")
            print(f"[DEBUG] Saving to file: {ADMIN_COOKIES_FILE}")
            
            if save_cookies_config(admin_cookies, ADMIN_COOKIES_FILE):
                print(f"[DEBUG] Successfully saved to {ADMIN_COOKIES_FILE}")
                st.success(f"Global cookies for {st.session_state.current_env} saved successfully!")
                
                # Update admin's session state to reflect the saved cookies
                st.session_state.cookies_config[st.session_state.current_env] = cookies_string
                print(f"[DEBUG] Updated admin session state cookies")
                
                # Update admin's user data as well
                if st.session_state.active_user in st.session_state.logged_in_users:
                    st.session_state.logged_in_users[st.session_state.active_user]['cookies_config'][st.session_state.current_env] = cookies_string
                    print(f"[DEBUG] Updated admin user data cookies")
                
                # Clear all users' cookies for this environment so they use admin cookies by default
                for username in st.session_state.logged_in_users:
                    if username != "adminadmin":  # Don't clear admin's own cookies
                        user_data = st.session_state.logged_in_users[username]
                        # Clear this environment's cookies for the user
                        user_data['cookies_config'][st.session_state.current_env] = ""
                        
                        # Save user's updated cookies config to file
                        user_cookies_file = user_data['file_paths']["COOKIES_CONFIG_FILE"]
                        save_cookies_config(user_data['cookies_config'], user_cookies_file)
                        print(f"[DEBUG] Cleared cookies for user: {username}")
                
                st.info("All users' cookies have been cleared to use the new global cookies.")
                st.rerun()  # Refresh the UI to show updated values
            else:
                print(f"[DEBUG] Failed to save to {ADMIN_COOKIES_FILE}")
                st.error("Failed to save global cookies")
        
        # Reset to defaults button for admin
        if st.button("Reset to Environment Default"):
            environments = load_environments_config()
            default_cookies = environments.get(st.session_state.current_env, {}).get('default_cookies', '')
            admin_cookies[st.session_state.current_env] = default_cookies
            
            if save_cookies_config(admin_cookies, ADMIN_COOKIES_FILE):
                # Update admin's session state to reflect the reset cookies
                st.session_state.cookies_config[st.session_state.current_env] = default_cookies
                
                # Update admin's user data as well
                if st.session_state.active_user in st.session_state.logged_in_users:
                    st.session_state.logged_in_users[st.session_state.active_user]['cookies_config'][st.session_state.current_env] = default_cookies
                
                st.success(f"Global cookies for {st.session_state.current_env} reset to environment default")
                st.rerun()
            else:
                st.error("Failed to reset global cookies")
                
    else:
        # Regular user - manage personal cookies (default empty)
        # Get current environment cookies string - default to empty for users
        env_cookies_string = st.session_state.cookies_config.get(st.session_state.current_env, "")

        # Show admin cookies for reference
        admin_cookies = {}
        if os.path.exists(ADMIN_COOKIES_FILE):
            try:
                admin_cookies = load_api_configs(ADMIN_COOKIES_FILE)
            except:
                admin_cookies = {}
        
        admin_cookie_value = admin_cookies.get(st.session_state.current_env, "")
        if admin_cookie_value:
            st.info(f"**Using Admin Cookie for {st.session_state.current_env}**")
            # st.text(admin_cookie_value)
            st.write("*Leave your cookies empty to use admin cookies automatically*")

        # Display as editable text area with string format
        cookies_string = st.text_area(
            "Your custom cookies",
            value=env_cookies_string,
            height=150,
            help="Leave empty to use admin cookies. Enter cookies in standard format: name1=value1; name2=value2",
            placeholder="Leave empty to use admin cookies..."
        )

        # Save button
        if st.button("Save Cookies"):
            # Save the cookies string (can be empty)
            st.session_state.cookies_config[st.session_state.current_env] = cookies_string

            # Save to file and update user data
            if save_cookies_config(st.session_state.cookies_config, file_path):
                _save_current_user_data()  # Update the logged_in_users dict
                if cookies_string.strip():
                    st.success(f"Custom cookies for {st.session_state.current_env} saved successfully!")
                else:
                    st.success(f"Cookies cleared for {st.session_state.current_env}. Will use admin cookies.")
            else:
                st.error("Failed to save cookies configuration")

        # Clear cookies button
        if st.button("Clear My Cookies"):
            # Set to empty string
            st.session_state.cookies_config[st.session_state.current_env] = ""
            
            # Save to file and update user data
            if save_cookies_config(st.session_state.cookies_config, file_path):
                _save_current_user_data()
                st.success(f"Cookies cleared for {st.session_state.current_env}. Will use admin cookies.")
            else:
                st.error("Failed to clear cookies")
            st.rerun()

        # Reset to admin defaults button (this now clears user cookies)
        if st.button("Use Admin Cookies"):
            # Set to empty string to use admin cookies
            st.session_state.cookies_config[st.session_state.current_env] = ""
            
            # Save to file and update user data
            save_cookies_config(st.session_state.cookies_config, file_path)
            _save_current_user_data()
            st.success(f"Will now use admin cookies for {st.session_state.current_env}")
            st.rerun()
def add_new_api():
    """Add new API."""
    with st.form("add_api_form"):
        api_name = st.text_input("API Name")
        
        # Instead of full URL, just ask for the path (with leading slash)
        api_path = st.text_input(
            "API Path (starts with /)", 
            help="Enter the API path starting with a slash (e.g., /subjectcomponent/list)"
        )

        method = st.selectbox("HTTP Method", ["GET", "POST", "PUT", "DELETE", "PATCH"])
        
        # Module selection dropdown
        module = st.selectbox(
            "Module",
            ["EX", "AD"],
            index=0,  # Default to EX
            help="EX: Assessment Module, AD: Administration Module"
        )

        submitted = st.form_submit_button("Add API")
        if submitted and api_name and api_path:
            # Ensure path starts with a slash
            api_path = ensure_path_format(api_path)

            # Combine base URL with path using selected module
            api_url = f"{get_current_base_url(st.session_state.current_env, module)}{api_path}"

            st.session_state.apis[api_name] = {
                "url": api_url,
                "path": api_path,  # Store the path separately
                "method": method,
                "module": module,  # Store the selected module
                "headers": {},
                "params": {},
                "body": {}
            }
            st.session_state.current_api = api_name

            # Save to file and update user data
            save_user_apis(st.session_state.apis, st.session_state.file_paths["USER_APIS_FILE"])
            _save_current_user_data()

            st.success(f"API '{api_name}' added successfully!")
            st.rerun()


def load_predefined_api(file_path):
    """Load predefined API for viewing without saving to user list"""
    # Load predefined API tests from JSON file
    predefined_configs = load_api_configs(file_path)

    if not predefined_configs:
        st.warning("No predefined API configurations found")
        return

    # Get current user's selected module
    current_module = st.session_state.get('selected_module', 'EX')
    
    # Filter predefined APIs to only show those matching current module
    filtered_configs = {
        api_name: config for api_name, config in predefined_configs.items()
        if config.get('module', 'EX') == current_module
    }
    
    if not filtered_configs:
        module_name = "Assessment" if current_module == "EX" else "Administration"
        st.warning(f"No predefined API configurations found for {module_name} ({current_module}) module")
        return

    # Create a dropdown for predefined APIs
    if filtered_configs:
        # Create a display list that shows module information
        api_display_options = []
        api_name_mapping = {}
        
        for api_name, config in filtered_configs.items():
            module = config.get('module', 'EX')
            module_name = "Assessment" if module == "EX" else "Administration"
            display_name = f"{api_name} ({module_name})"
            api_display_options.append(display_name)
            api_name_mapping[display_name] = api_name
        
        selected_display = st.selectbox(
            "Select Predefined API Test",
            api_display_options
        )
        
        selected_api = api_name_mapping[selected_display]
    else:
        selected_api = None

    if selected_api and st.button("Load Predefined Test"):
        # Convert the configuration to a complete API configuration
        api_config = filtered_configs[selected_api].copy()

        # Get the path from the config
        path = api_config.get("path", "")
        if not path:
            # If no path field, try url_path for backward compatibility
            path = api_config.get("url_path", "")

        # Ensure path starts with a slash
        path = ensure_path_format(path) if path else ""
        
        # Get module from config or default to EX
        config_module = api_config.get("module", "EX")
        
        # Special handling for Timer Job APIs - don't build URL here, will be built dynamically
        # Store the template path for Timer Job APIs
        if "{timer_job_id}" in path:
            # Keep the template path as is, URL will be built when sending request
            api_config["url_path"] = path
            api_config["path"] = path
            # Ensure timer_job_id is copied from predefined config
            if "timer_job_id" not in api_config and "timer_job_id" in filtered_configs[selected_api]:
                api_config["timer_job_id"] = filtered_configs[selected_api]["timer_job_id"]
            # Don't set api_config["url"] here - will be built dynamically in _handle_send_button
        else:
            # Add the full URL based on current environment and config module for non-timer APIs
            api_config["url"] = f"{get_current_base_url(st.session_state.current_env, config_module)}{path}"

            # Make sure path is stored
            api_config["path"] = path
        
        # Store the module from config
        api_config["module"] = config_module

        # Create a temporary name with special marker
        temp_api_name = f"temp_{selected_api}"

        # Store in session state directly in the apis dictionary, but with a flag
        st.session_state.apis[temp_api_name] = api_config
        st.session_state.apis[temp_api_name]["is_temporary"] = True
        st.session_state.apis[temp_api_name]["original_name"] = selected_api

        # Set as current API for viewing
        st.session_state.current_api = temp_api_name
        
        # Automatically switch to the correct module if different from current selection
        # Determine default module based on username if not set
        username = st.session_state.get('username', '')
        if username.upper() == "QA":
            default_module = "AD"
        else:
            default_module = "EX"
        
        current_selected_module = st.session_state.get('selected_module', default_module)
        if config_module != current_selected_module:
            st.session_state.selected_module = config_module
            st.info(f"Switched to {config_module} module to display the loaded API.")

        module_name = "Assessment" if config_module == "EX" else "Administration"
        st.success(f"Loaded '{selected_api}' API test (preview) - Module: {module_name} ({config_module})")
        st.info("This API is in preview mode. Save it to add permanently to your saved APIs.")
        st.rerun()

def _render_headers_section(api):
    """Render the headers section of the API tester"""
    with st.expander("Headers", expanded=False):
        st.write("Add or modify headers:")

        # Show current cookie information
        current_cookies = api.get('cookies', {})
        if current_cookies:
            st.info("**Current Cookies:**")
            cookie_display = []
            for name, value in current_cookies.items():
                # Truncate long cookie values for display
                display_value = value if len(value) <= 50 else value[:47] + "..."
                cookie_display.append(f"• {name} = {display_value}")
            
            if cookie_display:
                st.text("\n".join(cookie_display))
                
                # Add Cookie header if not already present
                cookie_header_value = "; ".join([f"{name}={value}" for name, value in current_cookies.items()])
                if 'headers' not in api:
                    api['headers'] = {}
                api['headers']['Cookie'] = cookie_header_value
        else:
            st.info("No cookies configured for this request")

        # Display existing headers
        if api.get('headers'):
            for header_key in list(api['headers'].keys()):
                col1, col2, col3 = st.columns([2, 3, 1])
                with col1:
                    st.text(header_key)
                with col2:
                    # Special handling for Cookie header to show it's auto-generated
                    if header_key.lower() == 'cookie':
                        st.text_input(f"Value for {header_key}", 
                                     value=api['headers'][header_key], 
                                     key=f"header_value_{header_key}",
                                     disabled=True,
                                     help="Auto-generated from cookie configuration")
                    else:
                        new_value = st.text_input(f"Value for {header_key}", 
                                                 value=api['headers'][header_key], 
                                                 key=f"header_value_{header_key}")
                        api['headers'][header_key] = new_value
                with col3:
                    # Don't allow deletion of auto-generated Cookie header
                    if header_key.lower() != 'cookie':
                        if st.button("Delete", key=f"delete_header_{header_key}"):
                            del api['headers'][header_key]
                            st.rerun()

        # Add new header
        new_header_key = st.text_input("New Header Name")
        new_header_value = st.text_input("New Header Value")
        if st.button("Add Header") and new_header_key:
            if 'headers' not in api:
                api['headers'] = {}
            api['headers'][new_header_key] = new_header_value
            st.rerun()

def _render_parameters_section(api):
    """Render the parameters section for GET requests"""
    with st.expander("Query Parameters", expanded=True):
        st.write("Add or modify query parameters:")

        # Display existing parameters
        if api.get('params'):
            for param_key in list(api['params'].keys()):
                col1, col2, col3 = st.columns([2, 3, 1])
                with col1:
                    st.text(param_key)
                with col2:
                    new_value = st.text_input(f"Value for {param_key}", 
                                            value=api['params'][param_key], 
                                            key=f"param_value_{param_key}")
                    api['params'][param_key] = new_value
                with col3:
                    if st.button("Delete", key=f"delete_param_{param_key}"):
                        del api['params'][param_key]
                        st.rerun()

        # Add new parameter
        new_param_key = st.text_input("New Parameter Name")
        new_param_value = st.text_input("New Parameter Value")
        if st.button("Add Parameter") and new_param_key:
            if 'params' not in api:
                api['params'] = {}
            api['params'][new_param_key] = new_param_value
            st.rerun()


def _generate_excel_template():
    """Generate Excel template with sample data for student upload"""
    try:
        import io
        
        # Create sample data with proper format
        template_data = {
            'StudentID': [
                'STU001',
                'STU002', 
                'STU003',
                'STU004',
                'STU005'
            ],
            'FutureStage': [
                1,
                2,
                1,
                3,
                2
            ],
            'FutureCourseVersionCode': [
                'CS101V1',
                'MATH201V2',
                'ENG101V1',
                'PHY301V3',
                'CHEM201V2'
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the template data
            df.to_excel(writer, sheet_name='StudentData', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['StudentData']
            
            # Add some formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Format header row
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet
            instructions_data = {
                'Column Name': ['StudentID', 'FutureStage', 'FutureCourseVersionCode'],
                'Description': [
                    'Unique identifier for the student (e.g., STU001, STU002)',
                    'Future stage number (integer: 1, 2, 3, etc.)',
                    'Course version code (e.g., CS101V1, MATH201V2)'
                ],
                'Data Type': ['Text', 'Number', 'Text'],
                'Required': ['Yes', 'Yes', 'Yes'],
                'Example': ['STU001', '1', 'CS101V1']
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Format instructions sheet
            instructions_worksheet = writer.sheets['Instructions']
            for col_num, column_title in enumerate(instructions_df.columns, 1):
                cell = instructions_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths for instructions
            for column in instructions_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                instructions_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error generating template: {str(e)}")
        return None


def _generate_excel_template_ex():
    """Generate Excel template with sample data for EX module API (Assessment Student Info V2)"""
    try:
        import io
        
        # Create sample data with proper format for EX module
        template_data = {
            'SubjectCode': [
                'MATH101',
                'ENG102', 
                'CS201',
                'PHY301',
                'CHEM205'
            ],
            'CourseCode': [
                'COURSE001',
                'COURSE002',
                'COURSE001',
                'COURSE003',
                'COURSE002'
            ],
            'SemesterId': [
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6'
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the template data
            df.to_excel(writer, sheet_name='AssessmentData', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['AssessmentData']
            
            # Add some formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Format header row
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet
            instructions_data = {
                'Column Name': ['SubjectCode', 'CourseCode', 'SemesterId'],
                'Description': [
                    'Subject code identifier (e.g., MATH101, ENG102)',
                    'Course code identifier (e.g., COURSE001, COURSE002)',
                    'Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)'
                ],
                'Data Type': ['Text', 'Text', 'Text (UUID)'],
                'Required': ['Yes', 'Yes', 'Yes'],
                'Example': ['MATH101', 'COURSE001', '3fa85f64-5717-4562-b3fc-2c963f66afa6']
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Format instructions sheet
            instructions_worksheet = writer.sheets['Instructions']
            for col_num, column_title in enumerate(instructions_df.columns, 1):
                cell = instructions_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths for instructions
            for column in instructions_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                instructions_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error generating EX template: {str(e)}")
        return None


def _generate_excel_template_student_subject():
    """Generate Excel template with sample data for Student Subject API (DEVAddStudentV2)"""
    try:
        import pandas as pd
        import io
        
        # Create sample data with proper format for Student Subject API
        template_data = {
            'SubjectCode': [
                'MATH101',
                'ENG102', 
                'CS201',
                'PHY301',
                'CHEM205'
            ],
            'StudentId': [
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '4fb96g75-6828-5673-c4gd-3d074g77bgb7',
                '5gc07h86-7939-6784-d5he-4e185h88cha8',
                '6hd18i97-8a4a-7895-e6if-5f296i99dib9',
                '7ie29j08-9b5b-8906-f7jg-6g307j00ejc0'
            ],
            'IsDrop': [
                'false',
                'false',
                'true',
                'false',
                'true'
            ],
            'SemesterId': [
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6'
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the template data
            df.to_excel(writer, sheet_name='StudentSubjectData', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['StudentSubjectData']
            
            # Add some formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Format header row
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet
            instructions_data = {
                'Column Name': ['SubjectCode', 'StudentId', 'IsDrop', 'SemesterId'],
                'Description': [
                    'Subject code identifier (e.g., MATH101, ENG102)',
                    'Student UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)',
                    'Drop status as string: "true" or "false" (default: "false")',
                    'Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)'
                ],
                'Data Type': ['Text', 'Text (UUID)', 'Text (true/false)', 'Text (UUID)'],
                'Required': ['Yes', 'Yes', 'Yes', 'Yes'],
                'Example': ['MATH101', '3fa85f64-5717-4562-b3fc-2c963f66afa6', 'false', '3fa85f64-5717-4562-b3fc-2c963f66afa6']
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Format instructions sheet
            instructions_worksheet = writer.sheets['Instructions']
            for col_num, column_title in enumerate(instructions_df.columns, 1):
                cell = instructions_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths for instructions
            for column in instructions_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                instructions_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error generating Student Subject template: {str(e)}")
        return None


def _generate_excel_template_course_student():
    """Generate Excel template with sample data for Course Student API (AssessmentStudentInfo/DEVAddStudentV2)"""
    try:
        import pandas as pd
        import io
        
        # Create sample data with proper format for Course Student API
        template_data = {
            'StudentId': [
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '4fb96g75-6828-5673-c4gd-3d074g77bgb7',
                '5gc07h86-7939-6784-d5he-4e185h88cha8',
                '6hd18i97-8a4a-7895-e6if-5f296i99dib9',
                '7ie29j08-9b5b-8906-f7jg-6g307j00ejc0'
            ],
            'CourseCode': [
                'COURSE001',
                'COURSE001',
                'COURSE002',
                'COURSE002',
                'COURSE003'
            ],
            'SemesterId': [
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6'
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the template data
            df.to_excel(writer, sheet_name='CourseStudentData', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['CourseStudentData']
            
            # Add some formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Format header row
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet
            instructions_data = {
                'Column Name': ['StudentId', 'CourseCode', 'SemesterId'],
                'Description': [
                    'Student UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)',
                    'Course code identifier (e.g., COURSE001, MATH101)',
                    'Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)'
                ],
                'Data Type': ['Text (UUID)', 'Text', 'Text (UUID)'],
                'Required': ['Yes', 'Yes', 'Yes'],
                'Example': ['3fa85f64-5717-4562-b3fc-2c963f66afa6', 'COURSE001', '3fa85f64-5717-4562-b3fc-2c963f66afa6']
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Format instructions sheet
            instructions_worksheet = writer.sheets['Instructions']
            for col_num, column_title in enumerate(instructions_df.columns, 1):
                cell = instructions_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths for instructions
            for column in instructions_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                instructions_worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add API information sheet
            api_info_data = {
                'API Information': [
                    'Endpoint', 'Method', 'Purpose', 'Input Format', 'Sample CourseCode'
                ],
                'Details': [
                    '/AssessmentStudentInfo/DEVAddStudentV2',
                    'POST',
                    'Add multiple students to a specific course in a semester',
                    'Excel file with StudentId, CourseCode, SemesterId columns',
                    'COURSE001, MATH101, ENG102, etc.'
                ]
            }
            
            api_info_df = pd.DataFrame(api_info_data)
            api_info_df.to_excel(writer, sheet_name='APIInfo', index=False)
            
            # Format API info sheet
            api_info_worksheet = writer.sheets['APIInfo']
            for col_num, column_title in enumerate(api_info_df.columns, 1):
                cell = api_info_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths for API info
            for column in api_info_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                api_info_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error generating Course Student template: {str(e)}")
        return None


def _render_excel_upload_section_course_student(api_name, api, file_paths):
    """Render Excel upload section for Course Student API (AssessmentStudentInfo/DEVAddStudentV2)"""
    
    st.subheader("📊 Excel Upload for Course Student Data")
    st.info("Upload an Excel file with columns: StudentId, CourseCode, SemesterId")
    
    # Template download section
    with st.expander("📥 Download Excel Template", expanded=True):
        st.write("**Get started with the correct format:**")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            - **StudentId**: Student UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)
            - **CourseCode**: Course code identifier (e.g., COURSE001, MATH101)
            - **SemesterId**: Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)
            """)
        
        with col2:
            template_data = _generate_excel_template_course_student()
            if template_data:
                st.download_button(
                    label="Download Template",
                    data=template_data,
                    file_name="course_student_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_template_course_{api_name}",
                    help="Download Excel template with sample data and instructions",
                    use_container_width=True
                )
    
    st.markdown("---")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        key=f"excel_upload_course_{api_name}",
        help="Upload Excel file with StudentId, CourseCode, SemesterId columns"
    )
    
    if uploaded_file is None:
        st.info("💡 **Tip**: Download the template above to get started with the correct format!")
    
    # Process uploaded file
    if uploaded_file is not None:
        try:
            # Create upload_data directory if it doesn't exist
            upload_dir = os.path.join(os.path.dirname(__file__), "upload_data")
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            
            # Generate unique filename with timestamp
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            username = st.session_state.get('username', 'unknown')
            
            # Clean names for file path
            clean_username = ''.join(c for c in username if c.isalnum() or c in '-_').strip()
            if not clean_username:
                clean_username = 'unknown'
            
            clean_api_name = ''.join(c for c in api_name if c.isalnum() or c in '-_').strip()
            if not clean_api_name:
                clean_api_name = 'course_api'
            
            file_extension = os.path.splitext(uploaded_file.name)[1]
            if not file_extension:
                file_extension = '.xlsx'
            
            saved_filename = f"{clean_username}_{clean_api_name}_{timestamp}{file_extension}"
            saved_file_path = os.path.join(upload_dir, saved_filename)
            
            # Save the uploaded file
            with open(saved_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Read Excel file
            df = pd.read_excel(uploaded_file)
            
            # Validate required columns
            required_columns = ['StudentId', 'CourseCode', 'SemesterId']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"Missing required columns: {', '.join(missing_columns)}")
                st.info(f"Required columns: {', '.join(required_columns)}")
            else:
                # Show preview of data
                st.success(f"✅ File uploaded successfully! Found {len(df)} course student records.")
                
                with st.expander("📋 Data Preview", expanded=True):
                    st.dataframe(df.head(10))
                    if len(df) > 10:
                        st.info(f"Showing first 10 rows of {len(df)} total rows")
                
                # Process data and fill JSON
                student_ids = []
                course_code = None
                semester_id = None
                
                for _, row in df.iterrows():
                    student_ids.append(str(row['StudentId']))
                    
                    # Use the first row's CourseCode and SemesterId for the entire dataset
                    if course_code is None:
                        course_code = str(row['CourseCode'])
                    if semester_id is None:
                        semester_id = str(row['SemesterId'])
                
                # Remove duplicates while preserving order
                student_ids = list(dict.fromkeys(student_ids))
                
                # Create the JSON structure for Course Student API
                json_body = {
                    "semesterId": semester_id,
                    "courseCode": course_code,
                    "studentIds": student_ids
                }
                
                # Update API body
                api['body'] = json_body
                
                # Update session state
                formatted_json = json.dumps(json_body, indent=2, ensure_ascii=False)
                body_json_key = f"original_body_json_{api_name}"
                st.session_state[body_json_key] = formatted_json
                st.session_state[f"json_body_course_{api_name}"] = formatted_json
                
                # Only update current user data in memory
                _save_current_user_data()
                
                # Show summary
                unique_students = len(student_ids)
                unique_courses = len(set(str(row['CourseCode']) for _, row in df.iterrows()))
                
                st.success(f"✅ Processed {len(df)} records into {unique_students} unique students!")
                st.info(f"📊 Summary: {unique_students} students for {unique_courses} course(s)")
                st.info(f"📋 CourseCode: {course_code}, SemesterId: {semester_id}")
                    
        except Exception as e:
            st.error(f"Error processing Excel file: {str(e)}")
            st.info("Please ensure the file format is correct and contains the required columns.")
    
    # JSON Body section (always show, with or without Excel upload)
    username = st.session_state.get('username', '')
    is_qa_account = username.upper().startswith("QA")
    json_body_expanded = not is_qa_account

    with st.expander("📝 JSON Body", expanded=json_body_expanded):
        # Show as JSON editor with better formatting
        if 'body' not in api:
            api['body'] = {
                "semesterId": "3fa85f64-5717-4562-b3fc-2c963f66afa6",
                "courseCode": "string",
                "studentIds": ["3fa85f64-5717-4562-b3fc-2c963f66afa6"]
            }
            
        # Keep track of original JSON to detect changes
        original_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
        body_json_key = f"original_body_json_{api_name}"
        
        if body_json_key not in st.session_state:
            st.session_state[body_json_key] = original_json
        
        # Add helpful buttons for common JSON operations
        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("Format JSON", key=f"format_json_course_{api_name}"):
                try:
                    # Parse and reformat the current JSON
                    current_json = st.session_state.get(f"json_body_course_{api_name}", original_json)
                    parsed = json.loads(current_json)
                    formatted_json = json.dumps(parsed, indent=2, ensure_ascii=False)
                    
                    # Update the API body and session state
                    api['body'] = parsed
                    st.session_state[f"json_body_course_{api_name}"] = formatted_json
                    st.session_state[body_json_key] = formatted_json
                    
                    # Only update current user data in memory
                    _save_current_user_data()
                    
                    st.success("✅ JSON formatted!")
                    st.rerun()
                except json.JSONDecodeError:
                    st.error("❌ Cannot format invalid JSON")
        
        with col2:
            st.info("💡 This API adds students to courses")

        body_json = st.text_area(
            "JSON Body (Auto-filled from Excel or manual entry)", 
            value=original_json,
            height=400,
            key=f"json_body_course_{api_name}",
            help="JSON body is auto-filled when Excel file is processed, or you can edit manually"
        )

        try:
            # Parse and validate the JSON
            parsed_body = json.loads(body_json)
            api['body'] = parsed_body
            
            # Only update session state if JSON changed
            if body_json != st.session_state.get(body_json_key, ""):
                st.session_state[body_json_key] = body_json
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
        except json.JSONDecodeError as e:
            st.error(f"❌ Invalid JSON format: {str(e)}")
            st.warning("⚠️ Changes will not be auto-saved until JSON is valid")
            
        # Show a preview of the parsed JSON structure if valid
        try:
            parsed_preview = json.loads(body_json)
            if parsed_preview:
                with st.expander("📋 JSON Structure Preview", expanded=False):
                    st.json(parsed_preview)
        except json.JSONDecodeError:
            pass


def _generate_excel_template_allocate_student():
    """Generate Excel template with sample data for Allocate Student API (DEVAllocateStudent dual API)"""
    try:
        import pandas as pd
        import io
        
        # Create sample data with proper format for Allocate Student API
        template_data = {
            'SubjectCode': [
                'MATH101',
                'MATH102', 
                'ENG201',
                'CS301',
                'CHEM205'
            ],
            'StudentId': [
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '4fb96g75-6828-5673-c4gd-3d074g77bgb7',
                '5gc07h86-7939-6784-d5he-4e185h88cha8',
                '6hd18i97-8a4a-7895-e6if-5f296i99dib9',
                '7ie29j08-9b5b-8906-f7jg-6g307j00ejc0'
            ],
            'IsDrop': [
                'false',
                'false',
                'true',
                'false',
                'true'
            ],
            'SemesterId': [
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                '3fa85f64-5717-4562-b3fc-2c963f66afa6'
            ],
            'CourseCode': [
                'MATH_COURSE',
                'MATH_COURSE',
                'ENGLISH_COURSE',
                'CS_COURSE',
                'CHEM_COURSE'
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the template data
            df.to_excel(writer, sheet_name='AllocateStudentData', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['AllocateStudentData']
            
            # Add some formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Format header row
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet
            instructions_data = {
                'Column Name': ['SubjectCode', 'StudentId', 'IsDrop', 'SemesterId', 'CourseCode'],
                'Description': [
                    'Subject code identifier (e.g., MATH101, ENG102)',
                    'Student UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)',
                    'Drop status as string: "true" or "false" (default: "false")',
                    'Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)',
                    'Course code identifier (e.g., COURSE001, can be same as SubjectCode)'
                ],
                'Data Type': ['Text', 'Text (UUID)', 'Text (true/false)', 'Text (UUID)', 'Text'],
                'Required': ['Yes', 'Yes', 'Yes', 'Yes', 'Yes'],
                'Example': ['MATH101', '3fa85f64-5717-4562-b3fc-2c963f66afa6', 'false', '3fa85f64-5717-4562-b3fc-2c963f66afa6', 'COURSE001']
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Format instructions sheet
            instructions_worksheet = writer.sheets['Instructions']
            for col_num, column_title in enumerate(instructions_df.columns, 1):
                cell = instructions_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths for instructions
            for column in instructions_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                instructions_worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add dual API explanation sheet
            explanation_data = {
                'Step': ['1', '2'],
                'API Called': [
                    'Add Real Student To Course Info (/AssessmentStudentInfo/DEVAddStudentV2)',
                    'Add Real Student to Subject Info (/AssessmentSubjectStudent/DEVAddStudentV2)'
                ],
                'Purpose': [
                    'Adds students to courses using CourseCode and StudentId',
                    'Adds students to subjects using SubjectCode, StudentId, and IsDrop'
                ],
                'Data Used': [
                    'semesterId, courseCode (from CourseCode column), studentIds (from StudentId column)',
                    'semesterId, studentInfos (SubjectCode, StudentId, IsDrop from respective columns)'
                ]
            }
            
            explanation_df = pd.DataFrame(explanation_data)
            explanation_df.to_excel(writer, sheet_name='DualAPIExplanation', index=False)
            
            # Format explanation sheet
            explanation_worksheet = writer.sheets['DualAPIExplanation']
            for col_num, column_title in enumerate(explanation_df.columns, 1):
                cell = explanation_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths for explanation
            for column in explanation_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)  # Wider for explanations
                explanation_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error generating Allocate Student template: {str(e)}")
        return None


def _render_excel_upload_section(api_name, api, file_paths):
    """Render Excel upload section for AD module APIs (DEVEXUpdateStudentUser)"""
    
    st.subheader("📊 Excel Upload for Student Data")
    st.info("Upload an Excel file with columns: StudentID, FutureStage, FutureCourseVersionCode")
    
    # Template download section
    with st.expander("📥 Download Excel Template", expanded=True):
        st.write("**Get started with the correct format:**")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            - **StudentID**: Unique student identifier (e.g., STU001, STU002)
            - **FutureStage**: Stage number as integer (1, 2, 3, etc.)
            - **FutureCourseVersionCode**: Course version code (e.g., CS101V1, MATH201V2)
            """)
        
        with col2:
            template_data = _generate_excel_template()
            if template_data:
                st.download_button(
                    label="Download Template",
                    data=template_data,
                    file_name="student_data_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_template_file_{api_name}",
                    help="Download Excel template with sample data and instructions",
                    use_container_width=True
                )
    
    st.markdown("---")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        key=f"excel_upload_{api_name}",
        help="Upload Excel file with StudentID, FutureStage, FutureCourseVersionCode columns"
    )
    
    if uploaded_file is None:
        st.info("💡 **Tip**: Download the template above to get started with the correct format!")
    
    # Process uploaded file
    if uploaded_file is not None:
        try:
            # Create upload_data directory if it doesn't exist
            upload_dir = os.path.join(os.path.dirname(__file__), "upload_data")
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            
            # Generate unique filename with timestamp - clean up names to avoid path issues
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            username = st.session_state.get('username', 'unknown')
            
            # Clean username - remove any characters that might cause path issues
            clean_username = ''.join(c for c in username if c.isalnum() or c in '-_').strip()
            if not clean_username:
                clean_username = 'unknown'
            
            # Clean API name - remove any characters that might cause path issues
            clean_api_name = ''.join(c for c in api_name if c.isalnum() or c in '-_').strip()
            if not clean_api_name:
                clean_api_name = 'api'
            
            file_extension = os.path.splitext(uploaded_file.name)[1]
            if not file_extension:
                file_extension = '.xlsx'  # Default extension
            
            saved_filename = f"{clean_username}_{clean_api_name}_{timestamp}{file_extension}"
            saved_file_path = os.path.join(upload_dir, saved_filename)
            
            # Save the uploaded file
            with open(saved_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # st.success(f"📁 File saved as: {saved_filename}")
            
            # Read Excel file directly from uploaded file buffer, not from saved file
            df = pd.read_excel(uploaded_file)
            
            # Validate required columns
            required_columns = ['StudentID', 'FutureStage', 'FutureCourseVersionCode']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"Missing required columns: {', '.join(missing_columns)}")
                st.info(f"Required columns: {', '.join(required_columns)}")
            else:
                # Show preview of data
                st.success(f"✅ File uploaded successfully! Found {len(df)} student records.")
                
                with st.expander("📋 Data Preview", expanded=True):
                    st.dataframe(df.head(10))
                    if len(df) > 10:
                        st.info(f"Showing first 10 rows of {len(df)} total rows")
                
                # Automatically process data and fill JSON (no button needed)
                # Convert DataFrame to the required JSON format
                students_list = []
                
                for _, row in df.iterrows():
                    student_entry = {
                        "studentId": str(row['StudentID']),
                        "futureStage": int(row['FutureStage']),
                        "futureCourseVersionCode": str(row['FutureCourseVersionCode'])
                    }
                    students_list.append(student_entry)
                
                # Create the final JSON structure
                json_body = {
                    "students": students_list
                }
                
                # Update API body
                api['body'] = json_body
                
                # Update session state
                formatted_json = json.dumps(json_body, indent=2, ensure_ascii=False)
                body_json_key = f"original_body_json_{api_name}"
                st.session_state[body_json_key] = formatted_json
                st.session_state[f"json_body_{api_name}"] = formatted_json
                
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
                st.success(f"✅ Automatically processed {len(students_list)} student records and filled JSON body!")
                
                # OLD MANUAL BUTTON (commented out - no longer needed)
                # if st.button("🔄 Process Data & Fill JSON", key=f"process_excel_{api_name}"):
                #     # Convert DataFrame to the required JSON format
                #     students_list = []
                #     
                #     for _, row in df.iterrows():
                #         student_entry = {
                #             "studentId": str(row['StudentID']),
                #             "futureStage": int(row['FutureStage']),
                #             "futureCourseVersionCode": str(row['FutureCourseVersionCode'])
                #         }
                #         students_list.append(student_entry)
                #     
                #     # Create the final JSON structure
                #     json_body = {
                #         "students": students_list
                #     }
                #     
                #     # Update API body
                #     api['body'] = json_body
                #     
                #     # Update session state
                #     formatted_json = json.dumps(json_body, indent=2, ensure_ascii=False)
                #     body_json_key = f"original_body_json_{api_name}"
                #     st.session_state[body_json_key] = formatted_json
                #     st.session_state[f"json_body_{api_name}"] = formatted_json
                #     
                #     # Auto-save
                #     st.session_state.apis[api_name] = api.copy()
                #     save_user_apis(st.session_state.apis, file_paths["USER_APIS_FILE"])
                #     _save_current_user_data()
                #     
                #     st.success(f"✅ Processed {len(students_list)} student records and filled JSON body!")
                #     st.rerun()
                    
        except Exception as e:
            st.error(f"Error processing Excel file: {str(e)}")
            st.info("Please ensure the file format is correct and contains the required columns.")
            
            # Show helpful info for common issues
            if "No such file or directory" in str(e):
                st.warning("💡 **File access issue**: Please try uploading the file again.")
            elif "BadZipFile" in str(e) or "XLRDError" in str(e):
                st.warning("� **File format issue**: Please ensure you're uploading a valid Excel file (.xlsx or .xls).")
    

    
    # JSON Body section (always show, with or without Excel upload)
    # Determine if JSON Body should be expanded by default
    username = st.session_state.get('username', '')
    is_qa_account = username.upper().startswith("QA")
    # For QA accounts in AD module, default to closed; otherwise expanded
    json_body_expanded = not is_qa_account  # QA accounts get False (closed), others get True (expanded)

    with st.expander("📝 JSON Body", expanded=json_body_expanded):
        # Show as JSON editor with better formatting
        if 'body' not in api:
            api['body'] = {"students": []}
            
        # Keep track of original JSON to detect changes
        original_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
        body_json_key = f"original_body_json_{api_name}"
        
        if body_json_key not in st.session_state:
            st.session_state[body_json_key] = original_json
        
        # Add helpful buttons for common JSON operations
        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("Format JSON", key=f"format_json_{api_name}"):
                try:
                    # Parse and reformat the current JSON
                    current_json = st.session_state.get(f"json_body_{api_name}", original_json)
                    parsed = json.loads(current_json)
                    formatted_json = json.dumps(parsed, indent=2, ensure_ascii=False)
                    
                    # Update the API body and session state
                    api['body'] = parsed
                    st.session_state[f"json_body_{api_name}"] = formatted_json
                    st.session_state[body_json_key] = formatted_json
                    
                    # Only update current user data in memory (no file save)
                    _save_current_user_data()
                    
                    st.success("✅ JSON formatted!")
                    st.rerun()
                except json.JSONDecodeError:
                    st.error("❌ Cannot format invalid JSON")
        
        with col2:
            st.info("💡 Tip: Upload Excel file above or manually edit JSON. Changes tracked in memory.")

        body_json = st.text_area(
            "JSON Body (Auto-filled from Excel or manual entry)", 
            value=original_json,
            height=400,
            key=f"json_body_{api_name}",
            help="JSON body is auto-filled when Excel file is processed, or you can edit manually. Changes are tracked in memory and will be saved when you click 'Save API'."
        )

        try:
            # Parse the JSON
            parsed_body = json.loads(body_json)
            api['body'] = parsed_body
            
            # Show JSON validation status
            # st.success("✅ Valid JSON")
            
            # Show student count if applicable
            if isinstance(parsed_body, dict) and 'students' in parsed_body:
                student_count = len(parsed_body['students'])
                # st.info(f"📊 Contains {student_count} student record(s)")
            
            # Auto-save if JSON has changed and is valid
            if body_json != st.session_state[body_json_key]:
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
                # Update original JSON after saving
                st.session_state[body_json_key] = body_json
                
                # Show a subtle feedback that changes are tracked
                # st.info("💾 Changes tracked in memory")
            
        except json.JSONDecodeError as e:
            st.error(f"❌ Invalid JSON format: {str(e)}")
            st.warning("⚠️ Changes will not be auto-saved until JSON is valid")
            
        # Show a preview of the parsed JSON structure if valid
        try:
            parsed_preview = json.loads(body_json)
            if parsed_preview:
                with st.expander("📋 JSON Structure Preview", expanded=False):
                    st.json(parsed_preview)
        except json.JSONDecodeError:
            # Preview not available for invalid JSON
            pass


def _render_excel_upload_section_ex(api_name, api, file_paths):
    """Render Excel upload section for EX module APIs (DEVCreateDataV2)"""
    
    st.subheader("📊 Excel Upload for Assessment Student Data")
    st.info("Upload an Excel file with columns: SubjectCode, CourseCode, SemesterId")
    
    # Student size configuration
    with st.expander("⚙️ Student Size Configuration", expanded=True):
        col1, col2 = st.columns(2)
        
        # Get previous values to detect changes
        prev_student_size = st.session_state.get(f"student_size_value_{api_name}", 20)
        prev_max_student_size = st.session_state.get(f"max_student_size_value_{api_name}", 40)
        
        with col1:
            student_size = st.number_input(
                "Student Size",
                min_value=1,
                max_value=1000,
                value=prev_student_size,
                step=1,
                help="Number of students to generate (default: 20)",
                key=f"student_size_{api_name}"
            )
        
        with col2:
            max_student_size = st.number_input(
                "Max Student Size",
                min_value=1,
                max_value=1000,
                value=prev_max_student_size,
                step=1,
                help="Maximum number of students allowed (default: 40)",
                key=f"max_student_size_{api_name}"
            )
        
        # Auto-save when values change
        if student_size != prev_student_size or max_student_size != prev_max_student_size:
            # Store values in session state for use in JSON generation
            st.session_state[f"student_size_value_{api_name}"] = student_size
            st.session_state[f"max_student_size_value_{api_name}"] = max_student_size
            
            # Auto-update the JSON body if it exists
            if 'body' in api and isinstance(api['body'], dict):
                try:
                    # Update the API body directly
                    api['body']['studentSize'] = student_size
                    api['body']['maxStudentSize'] = max_student_size
                    
                    # Update the JSON in the text area as well
                    formatted_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
                    body_json_key = f"original_body_json_{api_name}"
                    st.session_state[body_json_key] = formatted_json
                    st.session_state[f"json_body_ex_{api_name}"] = formatted_json
                    
                    # Update current user data in memory
                    _save_current_user_data()
                    
                    st.success(f"✅ Auto-saved! StudentSize: {student_size}, MaxStudentSize: {max_student_size}")
                    st.rerun()
                except Exception as e:
                    st.warning(f"Could not auto-update JSON: {str(e)}")
        else:
            # Store values in session state for use in JSON generation
            st.session_state[f"student_size_value_{api_name}"] = student_size
            st.session_state[f"max_student_size_value_{api_name}"] = max_student_size
    
    # Template download section
    with st.expander("📥 Download Excel Template", expanded=True):
        st.write("**Get started with the correct format:**")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            - **SubjectCode**: Subject code identifier (e.g., MATH101, ENG102)
            - **CourseCode**: Course code identifier (e.g., COURSE001, COURSE002)
            - **SemesterId**: Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)
            """)
        
        with col2:
            template_data = _generate_excel_template_ex()
            if template_data:
                st.download_button(
                    label="Download Template",
                    data=template_data,
                    file_name="enroll_fake_student.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_template_ex_file_{api_name}",
                    help="Download Excel template with sample data and instructions",
                    use_container_width=True
                )
    
    st.markdown("---")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        key=f"excel_upload_ex_{api_name}",
        help="Upload Excel file with SubjectCode, CourseCode, SemesterId columns"
    )
    
    if uploaded_file is None:
        st.info("💡 **Tip**: Download the template above to get started with the correct format!")
    
    # Process uploaded file
    if uploaded_file is not None:
        try:
            # Create upload_data directory if it doesn't exist
            upload_dir = os.path.join(os.path.dirname(__file__), "upload_data")
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            
            # Generate unique filename with timestamp - clean up names to avoid path issues
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            username = st.session_state.get('username', 'unknown')
            
            # Clean username - remove any characters that might cause path issues
            clean_username = ''.join(c for c in username if c.isalnum() or c in '-_').strip()
            if not clean_username:
                clean_username = 'unknown'
            
            # Clean API name - remove any characters that might cause path issues
            clean_api_name = ''.join(c for c in api_name if c.isalnum() or c in '-_').strip()
            if not clean_api_name:
                clean_api_name = 'api'
            
            file_extension = os.path.splitext(uploaded_file.name)[1]
            if not file_extension:
                file_extension = '.xlsx'  # Default extension
            
            saved_filename = f"{clean_username}_{clean_api_name}_{timestamp}{file_extension}"
            saved_file_path = os.path.join(upload_dir, saved_filename)
            
            # Save the uploaded file
            with open(saved_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Read Excel file directly from uploaded file buffer, not from saved file
            df = pd.read_excel(uploaded_file)
            
            # Validate required columns
            required_columns = ['SubjectCode', 'CourseCode', 'SemesterId']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"Missing required columns: {', '.join(missing_columns)}")
                st.info(f"Required columns: {', '.join(required_columns)}")
            else:
                # Show preview of data
                st.success(f"✅ File uploaded successfully! Found {len(df)} assessment records.")
                
                with st.expander("📋 Data Preview", expanded=True):
                    st.dataframe(df.head(10))
                    if len(df) > 10:
                        st.info(f"Showing first 10 rows of {len(df)} total rows")
                
                # Automatically process data and fill JSON (no button needed)
                # Convert DataFrame to the required JSON format
                subject_codes_list = []
                course_code = None
                semester_id = None
                
                for _, row in df.iterrows():
                    subject_codes_list.append(str(row['SubjectCode']))
                    # Use the first row's CourseCode and SemesterId for the entire dataset
                    if course_code is None:
                        course_code = str(row['CourseCode'])
                    if semester_id is None:
                        semester_id = str(row['SemesterId'])
                
                # Remove duplicates while preserving order
                subject_codes_list = list(dict.fromkeys(subject_codes_list))
                
                # Get student size values from session state (with defaults)
                student_size = st.session_state.get(f"student_size_value_{api_name}", 20)
                max_student_size = st.session_state.get(f"max_student_size_value_{api_name}", 40)
                
                # Create the final JSON structure
                json_body = {
                    "maxStudentSize": max_student_size,
                    "studentSize": student_size,
                    "semesterId": semester_id,
                    "courseCode": course_code,
                    "subjectCodes": subject_codes_list
                }
                
                # Update API body
                api['body'] = json_body
                
                # Update session state
                formatted_json = json.dumps(json_body, indent=2, ensure_ascii=False)
                body_json_key = f"original_body_json_{api_name}"
                st.session_state[body_json_key] = formatted_json
                st.session_state[f"json_body_{api_name}"] = formatted_json
                
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
                st.success(f"✅ Automatically processed {len(subject_codes_list)} unique subject codes and filled JSON body!")
                st.info(f"📋 Using CourseCode: {course_code}, SemesterId: {semester_id}")
                    
        except Exception as e:
            st.error(f"Error processing Excel file: {str(e)}")
            st.info("Please ensure the file format is correct and contains the required columns.")
            
            # Show helpful info for common issues
            if "No such file or directory" in str(e):
                st.warning("💡 **File access issue**: Please try uploading the file again.")
            elif "BadZipFile" in str(e) or "XLRDError" in str(e):
                st.warning("⚠️ **File format issue**: Please ensure you're uploading a valid Excel file (.xlsx or .xls).")
    
    # JSON Body section (always show, with or without Excel upload)
    # Determine if JSON Body should be expanded by default
    username = st.session_state.get('username', '')
    is_qa_account = username.upper().startswith("QA")
    # For QA accounts in EX module, default to closed; otherwise expanded
    json_body_expanded = not is_qa_account  # QA accounts get False (closed), others get True (expanded)

    with st.expander("📝 JSON Body", expanded=json_body_expanded):
        # Show as JSON editor with better formatting
        if 'body' not in api:
            # Get student size values from session state (with defaults)
            student_size = st.session_state.get(f"student_size_value_{api_name}", 20)
            max_student_size = st.session_state.get(f"max_student_size_value_{api_name}", 40)
            
            api['body'] = {
                "maxStudentSize": max_student_size,
                "studentSize": student_size,
                "semesterId": "3fa85f64-5717-4562-b3fc-2c963f66afa6",
                "courseCode": "string",
                "subjectCodes": ["string"]
            }
            
        # Keep track of original JSON to detect changes
        original_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
        body_json_key = f"original_body_json_{api_name}"
        
        if body_json_key not in st.session_state:
            st.session_state[body_json_key] = original_json
        
        # Add helpful buttons for common JSON operations
        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("Format JSON", key=f"format_json_ex_{api_name}"):
                try:
                    # Parse and reformat the current JSON
                    current_json = st.session_state.get(f"json_body_{api_name}", original_json)
                    parsed = json.loads(current_json)
                    formatted_json = json.dumps(parsed, indent=2, ensure_ascii=False)
                    
                    # Update the API body and session state
                    api['body'] = parsed
                    st.session_state[f"json_body_{api_name}"] = formatted_json
                    st.session_state[body_json_key] = formatted_json
                    
                    # Only update current user data in memory (no file save)
                    _save_current_user_data()
                    
                    st.success("✅ JSON formatted!")
                    st.rerun()
                except json.JSONDecodeError:
                    st.error("❌ Cannot format invalid JSON")
        
        with col2:
            st.info("💡 Tip: Upload Excel file above or manually edit JSON. Changes tracked in memory.")

        body_json = st.text_area(
            "JSON Body (Auto-filled from Excel or manual entry)", 
            value=original_json,
            height=400,
            key=f"json_body_ex_{api_name}",
            help="JSON body is auto-filled when Excel file is processed, or you can edit manually. Changes are tracked in memory and will be saved when you click 'Save API'."
        )

        try:
            # Parse the JSON
            parsed_body = json.loads(body_json)
            api['body'] = parsed_body
            
            # Show subject codes count if applicable
            if isinstance(parsed_body, dict) and 'subjectCodes' in parsed_body:
                subject_count = len(parsed_body['subjectCodes'])
                # st.info(f"📊 Contains {subject_count} subject code(s)")
            
            # Auto-save if JSON has changed and is valid
            if body_json != st.session_state[body_json_key]:
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
                # Update original JSON after saving
                st.session_state[body_json_key] = body_json
                
        except json.JSONDecodeError as e:
            st.error(f"❌ Invalid JSON format: {str(e)}")
            st.warning("⚠️ Changes will not be auto-saved until JSON is valid")
            
        # Show a preview of the parsed JSON structure if valid
        try:
            parsed_preview = json.loads(body_json)
            if parsed_preview:
                with st.expander("📋 JSON Structure Preview", expanded=False):
                    st.json(parsed_preview)
        except json.JSONDecodeError:
            # Preview not available for invalid JSON
            pass


def _render_excel_upload_section_student_subject(api_name, api, file_paths):
    """Render Excel upload section for Student Subject API (DEVAddStudentV2)"""
    
    st.subheader("📊 Excel Upload for Student Subject Data")
    st.info("Upload an Excel file with columns: SubjectCode, StudentId, IsDrop, SemesterId")
    
    # Template download section
    with st.expander("📥 Download Excel Template", expanded=True):
        st.write("**Get started with the correct format:**")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            - **SubjectCode**: Subject code identifier (e.g., MATH101, ENG102)
            - **StudentId**: Student UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)
            - **IsDrop**: Drop status as string: "true" or "false" (default: "false")
            - **SemesterId**: Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)
            """)
        
        with col2:
            template_data = _generate_excel_template_student_subject()
            if template_data:
                st.download_button(
                    label="Download Template",
                    data=template_data,
                    file_name="enroll_real_student.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_template_student_subject_file_{api_name}",
                    help="Download Excel template with sample data and instructions",
                    use_container_width=True
                )
    
    st.markdown("---")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        key=f"excel_upload_student_subject_{api_name}",
        help="Upload Excel file with SubjectCode, StudentId, IsDrop, SemesterId columns"
    )
    
    if uploaded_file is None:
        st.info("💡 **Tip**: Download the template above to get started with the correct format!")
    
    # Process uploaded file
    if uploaded_file is not None:
        try:
            # Create upload_data directory if it doesn't exist
            upload_dir = os.path.join(os.path.dirname(__file__), "upload_data")
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            
            # Generate unique filename with timestamp - clean up names to avoid path issues
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            username = st.session_state.get('username', 'unknown')
            
            # Clean username - remove any characters that might cause path issues
            clean_username = ''.join(c for c in username if c.isalnum() or c in '-_').strip()
            if not clean_username:
                clean_username = 'unknown'
            
            # Clean API name - remove any characters that might cause path issues
            clean_api_name = ''.join(c for c in api_name if c.isalnum() or c in '-_').strip()
            if not clean_api_name:
                clean_api_name = 'api'
            
            file_extension = os.path.splitext(uploaded_file.name)[1]
            if not file_extension:
                file_extension = '.xlsx'  # Default extension
            
            saved_filename = f"{clean_username}_{clean_api_name}_{timestamp}{file_extension}"
            saved_file_path = os.path.join(upload_dir, saved_filename)
            
            # Save the uploaded file
            with open(saved_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Read Excel file directly from uploaded file buffer, not from saved file
            df = pd.read_excel(uploaded_file)
            
            # Validate required columns
            required_columns = ['SubjectCode', 'StudentId', 'IsDrop', 'SemesterId']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"Missing required columns: {', '.join(missing_columns)}")
                st.info(f"Required columns: {', '.join(required_columns)}")
            else:
                # Show preview of data
                st.success(f"✅ File uploaded successfully! Found {len(df)} student subject records.")
                
                with st.expander("📋 Data Preview", expanded=True):
                    st.dataframe(df.head(10))
                    if len(df) > 10:
                        st.info(f"Showing first 10 rows of {len(df)} total rows")
                
                # Automatically process data and fill JSON (no button needed)
                # Convert DataFrame to the required JSON format
                student_infos_list = []
                semester_id = None
                
                for _, row in df.iterrows():
                    # Convert IsDrop string to boolean
                    is_drop_str = str(row['IsDrop']).lower().strip()
                    is_drop = is_drop_str == 'true'
                    
                    student_info = {
                        "subjectCode": str(row['SubjectCode']),
                        "studentId": str(row['StudentId']),
                        "isDrop": is_drop
                    }
                    student_infos_list.append(student_info)
                    
                    # Use the first row's SemesterId for the entire dataset
                    if semester_id is None:
                        semester_id = str(row['SemesterId'])
                
                # Create the final JSON structure
                json_body = {
                    "semesterId": semester_id,
                    "studentInfos": student_infos_list
                }
                
                # Update API body
                api['body'] = json_body
                
                # Update session state
                formatted_json = json.dumps(json_body, indent=2, ensure_ascii=False)
                body_json_key = f"original_body_json_{api_name}"
                st.session_state[body_json_key] = formatted_json
                st.session_state[f"json_body_{api_name}"] = formatted_json
                
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
                st.success(f"✅ Automatically processed {len(student_infos_list)} student subject records and filled JSON body!")
                st.info(f"📋 Using SemesterId: {semester_id}")
                    
        except Exception as e:
            st.error(f"Error processing Excel file: {str(e)}")
            st.info("Please ensure the file format is correct and contains the required columns.")
            
            # Show helpful info for common issues
            if "No such file or directory" in str(e):
                st.warning("💡 **File access issue**: Please try uploading the file again.")
            elif "BadZipFile" in str(e) or "XLRDError" in str(e):
                st.warning("⚠️ **File format issue**: Please ensure you're uploading a valid Excel file (.xlsx or .xls).")
    
    # JSON Body section (always show, with or without Excel upload)
    # Determine if JSON Body should be expanded by default
    username = st.session_state.get('username', '')
    is_qa_account = username.upper().startswith("QA")
    # For QA accounts in EX module, default to closed; otherwise expanded
    json_body_expanded = not is_qa_account  # QA accounts get False (closed), others get True (expanded)

    with st.expander("📝 JSON Body", expanded=json_body_expanded):
        # Show as JSON editor with better formatting
        if 'body' not in api:
            api['body'] = {
                "semesterId": "3fa85f64-5717-4562-b3fc-2c963f66afa6",
                "studentInfos": [
                    {
                        "subjectCode": "string",
                        "studentId": "3fa85f64-5717-4562-b3fc-2c963f66afa6",
                        "isDrop": True
                    }
                ]
            }
            
        # Keep track of original JSON to detect changes
        original_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
        body_json_key = f"original_body_json_{api_name}"
        
        if body_json_key not in st.session_state:
            st.session_state[body_json_key] = original_json
        
        # Add helpful buttons for common JSON operations
        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("Format JSON", key=f"format_json_student_subject_{api_name}"):
                try:
                    # Parse and reformat the current JSON
                    current_json = st.session_state.get(f"json_body_{api_name}", original_json)
                    parsed = json.loads(current_json)
                    formatted_json = json.dumps(parsed, indent=2, ensure_ascii=False)
                    
                    # Update the API body and session state
                    api['body'] = parsed
                    st.session_state[f"json_body_{api_name}"] = formatted_json
                    st.session_state[body_json_key] = formatted_json
                    
                    # Only update current user data in memory (no file save)
                    _save_current_user_data()
                    
                    st.success("✅ JSON formatted!")
                    st.rerun()
                except json.JSONDecodeError:
                    st.error("❌ Cannot format invalid JSON")
        
        with col2:
            st.info("💡 Tip: Upload Excel file above or manually edit JSON. Changes tracked in memory.")

        body_json = st.text_area(
            "JSON Body (Auto-filled from Excel or manual entry)", 
            value=original_json,
            height=400,
            key=f"json_body_student_subject_{api_name}",
            help="JSON body is auto-filled when Excel file is processed, or you can edit manually. Changes are tracked in memory and will be saved when you click 'Save API'."
        )

        try:
            # Parse the JSON
            parsed_body = json.loads(body_json)
            api['body'] = parsed_body
            
            # Show student info count if applicable
            if isinstance(parsed_body, dict) and 'studentInfos' in parsed_body:
                student_count = len(parsed_body['studentInfos'])
                # st.info(f"📊 Contains {student_count} student info record(s)")
            
            # Auto-save if JSON has changed and is valid
            if body_json != st.session_state[body_json_key]:
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
                # Update original JSON after saving
                st.session_state[body_json_key] = body_json
                
        except json.JSONDecodeError as e:
            st.error(f"❌ Invalid JSON format: {str(e)}")
            st.warning("⚠️ Changes will not be auto-saved until JSON is valid")
            
        # Show a preview of the parsed JSON structure if valid
        try:
            parsed_preview = json.loads(body_json)
            if parsed_preview:
                with st.expander("📋 JSON Structure Preview", expanded=False):
                    st.json(parsed_preview)
        except json.JSONDecodeError:
            # Preview not available for invalid JSON
            pass


def _update_batch_mark_values(api_name, api, max_mark_value, min_mark_value):
    """Update batch processing mark values in session state and API body"""
    # Update session state for persistence
    st.session_state[f"max_mark_value_{api_name}"] = max_mark_value
    st.session_state[f"min_mark_value_{api_name}"] = min_mark_value
    
    # Update API body
    if 'body' not in api:
        api['body'] = {}
    
    api['body']['maxMark'] = int(max_mark_value)
    api['body']['minMark'] = int(min_mark_value)
    
    # Save to user data for persistence
    _save_current_user_data()


def _render_auto_mark_entry_section(api_name, api, file_paths):
    """Render special section for Auto Mark Entry API with batch processing"""
    
    st.subheader("📝 Auto Mark Entry API")
    st.info("Choose between Batch Processing (for multiple subjects) or Manual JSON Input (for single subject)")
    
    # Input mode selection
    input_mode = st.radio(
        "Input Mode",
        ["🚀 Batch Processing", "📝 Manual JSON Input"],
        key=f"input_mode_{api_name}",
        help="Batch Processing: Automatically call API for multiple Subject IDs | Manual JSON: Edit JSON body directly"
    )
    
    if input_mode == "🚀 Batch Processing":
        # Batch configuration section
        with st.expander("⚙️ Batch Configuration", expanded=True):
            st.write("**Configure batch parameters for Auto Mark Entry:**")
            
            # Semester ID input
            semester_id = st.text_input(
                "Semester ID",
                value=api.get('body', {}).get('semesterId', ''),
                key=f"semester_id_{api_name}",
                help="The semester ID for all subject entries"
            )
            
            # Max Mark input
            col1, col2 = st.columns(2)
            
            # Initialize session state for auto-save if not exists
            if f"max_mark_value_{api_name}" not in st.session_state:
                st.session_state[f"max_mark_value_{api_name}"] = api.get('body', {}).get('maxMark', 100)
            if f"min_mark_value_{api_name}" not in st.session_state:
                st.session_state[f"min_mark_value_{api_name}"] = api.get('body', {}).get('minMark', 0)
            
            with col1:
                max_mark = st.number_input(
                    "Max Mark",
                    min_value=0,
                    max_value=100,
                    value=st.session_state[f"max_mark_value_{api_name}"],
                    key=f"max_mark_{api_name}",
                    help="Maximum mark value",
                    on_change=lambda: _update_batch_mark_values(api_name, api, st.session_state[f"max_mark_{api_name}"], st.session_state[f"min_mark_{api_name}"])
                )
            
            with col2:
                min_mark = st.number_input(
                    "Min Mark",
                    min_value=0,
                    max_value=100,
                    value=st.session_state[f"min_mark_value_{api_name}"],
                    key=f"min_mark_{api_name}",
                    help="Minimum mark value",
                    on_change=lambda: _update_batch_mark_values(api_name, api, st.session_state[f"max_mark_{api_name}"], st.session_state[f"min_mark_{api_name}"])
                )
            
            st.markdown("---")
            
            # Subject Codes list input
            st.write("**Subject Codes List:**")
            st.write("Enter Subject Codes (one per line or comma-separated)")
            
            subject_ids_input = st.text_area(
                "Subject Codes",
                value="",
                height=150,
                key=f"subject_ids_{api_name}",
                help="Enter Subject Codes separated by new lines or commas. Example:\naaa\nbbb\nccc\nor: aaa, bbb, ccc",
                placeholder="Enter Subject Codes here...\nExample:\nB34_4\nB34_5"
            )
            
            # Parse subject IDs
            subject_ids = []
            if subject_ids_input:
                # Split by newlines and commas, then clean up
                raw_ids = subject_ids_input.replace(',', '\n').split('\n')
                subject_ids = [sid.strip() for sid in raw_ids if sid.strip()]
            
            # Display parsed IDs
            if subject_ids:
                st.success(f"✅ Found {len(subject_ids)} Subject Code(s)")
                with st.expander("📋 Preview Subject Codes", expanded=False):
                    for i, sid in enumerate(subject_ids, 1):
                        st.text(f"{i}. {sid}")
            else:
                st.warning("⚠️ No Subject Codes entered yet")
            
            st.markdown("---")
            
            # Subject Codes list input
            st.write("**Student IDs List:**")
            st.write("Enter Student IDs (one per line or comma-separated)")
            
            student_ids_input = st.text_area(
                "Subject IDs",
                value="",
                height=150,
                key=f"student_ids_{api_name}",
                help="Enter Student IDs separated by new lines or commas. Example:\naaa\nbbb\nccc\nor: aaa, bbb, ccc",
                placeholder="Enter StudentIds here...\nExample:\n3fa85f64-5717-4562-b3fc-2c963f66afa1\n3fa85f64-5717-4562-b3fc-2c963f66afa2"
            )
            
            # Parse subject IDs
            student_ids = []
            if student_ids_input:
                # Split by newlines and commas, then clean up
                raw_ids = student_ids_input.replace(',', '\n').split('\n')
                student_ids = [sid.strip() for sid in raw_ids if sid.strip()]
            
            # Display parsed IDs
            if student_ids:
                st.success(f"✅ Found {len(student_ids)} Subject Code(s)")
                with st.expander("📋 Preview Student IDs", expanded=False):
                    for i, sid in enumerate(student_ids, 1):
                        st.text(f"{i}. {sid}")
            else:
                st.warning("⚠️ No Student IDs entered yet")
            
            st.markdown("---")
            
            # Update API body with current values
            if 'body' not in api:
                api['body'] = {}
            
            api['body']['semesterId'] = semester_id
            api['body']['maxMark'] = int(max_mark)
            api['body']['minMark'] = int(min_mark)
            
            # Store subject IDs in session state for batch processing
            st.session_state[f'batch_subject_ids_{api_name}'] = subject_ids
            st.session_state[f'batch_student_ids_{api_name}'] = student_ids
        
        # Batch execution button
        if subject_ids and semester_id:
            st.markdown("---")
            
            # Explain what batch processing does
            with st.expander("ℹ️ How Batch Processing Works", expanded=False):
                st.write("**What does Batch Processing do?**")
                st.write("1. 📋 Takes your list of Subject IDs")
                st.write("2. 🔁 Loops through each Subject ID one by one")
                st.write("3. 📞 Calls the Auto Mark Entry API for each Subject ID")
                st.write("4. ✅ Uses the same Semester ID, Max Mark, and Min Mark for all calls")
                st.write("5. 📊 Shows progress and results in real-time")
                st.write("")
                st.write("**Example:**")
                st.code("""
If you have 3 Subject Codes: [aaa, bbb, ccc]
With Semester ID: xxx, Max Mark: 100, Min Mark: 60

It will make 3 API calls:
1. POST /automarkentry with {semesterId: xxx, subjectCode: aaa, maxMark: 100, minMark: 60}
2. POST /automarkentry with {semesterId: xxx, subjectCode: bbb, maxMark: 100, minMark: 60}
3. POST /automarkentry with {semesterId: xxx, subjectCode: ccc, maxMark: 100, minMark: 60}
                """)
            
            if st.button("🚀 Execute Batch Processing", type="primary", key=f"batch_execute_{api_name}"):
                _handle_auto_mark_entry_batch(api_name, api, subject_ids, student_ids, file_paths)
        else:
            st.info("💡 Enter Semester ID and at least one Subject Code to enable batch processing")
    
    else:  # Manual JSON Input mode
        with st.expander("📝 Manual JSON Input", expanded=True):
            st.write("**Edit the JSON body directly for single API execution:**")
            st.info("💡 Use this mode when you want to call the API once with specific parameters")
            
            # Show current body as JSON
            if 'body' not in api:
                api['body'] = {
                    'semesterId': '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                    'subjectCode': 'B34_4',
                    'maxMark': 100,
                    'minMark': 0
                }
            
            original_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
            
            # Add helpful buttons for common JSON operations
            col1, col2, col3 = st.columns(3)
            with col1:
                if st.button("📋 Reset to Default", key=f"reset_json_{api_name}"):
                    api['body'] = {
                        'semesterId': '3fa85f64-5717-4562-b3fc-2c963f66afa6',
                        'subjectCode': 'B34_4',
                        'maxMark': 100,
                        'minMark': 0
                    }
                    st.rerun()
            
            with col2:
                if st.button("🔄 Format JSON", key=f"format_json_{api_name}"):
                    try:
                        parsed = json.loads(st.session_state.get(f"json_body_manual_{api_name}", original_json))
                        formatted = json.dumps(parsed, indent=2, ensure_ascii=False)
                        st.session_state[f"json_body_manual_{api_name}"] = formatted
                        st.rerun()
                    except:
                        st.error("Cannot format invalid JSON")
            
            with col3:
                st.write("")  # Empty space for alignment
            
            body_json = st.text_area(
                "JSON Body",
                value=original_json,
                height=250,
                key=f"json_body_manual_{api_name}",
                help="Edit JSON body for single API execution. Make sure to include all required fields: semesterId, subjectId, maxMark, minMark"
            )
            
            try:
                parsed_body = json.loads(body_json)
                api['body'] = parsed_body
                
                # Validate required fields
                required_fields = ['semesterId', 'subjectId', 'maxMark', 'minMark']
                missing_fields = [field for field in required_fields if field not in parsed_body]
                
                if missing_fields:
                    st.warning(f"⚠️ Missing required fields: {', '.join(missing_fields)}")
                else:
                    st.success("✅ JSON is valid and contains all required fields")
                    
                    # Show a preview of what will be sent
                    with st.expander("📋 Request Preview", expanded=False):
                        st.json(parsed_body)
                        
            except json.JSONDecodeError as e:
                st.error(f"❌ Invalid JSON: {str(e)}")


def _handle_auto_mark_entry_batch(api_name, api, subject_ids, student_ids, file_paths):
    """Handle batch processing for Auto Mark Entry API"""
    
    total_ids = len(subject_ids)
    success_count = 0
    failed_count = 0
    results = []
    
    # Create progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    st.write("---")
    st.subheader("📊 Batch Processing Results")

    print("subject_ids:", subject_ids)  # Debugging line
    
    # Process each subject ID
    for i, subject_id in enumerate(subject_ids, 1):
        status_text.text(f"Processing {i}/{total_ids}: {subject_id}")
        progress_bar.progress(i / total_ids)
        
        # Create API call for this subject ID
        batch_api = api.copy()
        batch_api['body'] = batch_api.get('body', {}).copy()
        batch_api['body']['subjectCode'] = subject_id
        batch_api['body']['studentIds'] = student_ids
        
        try:
            # Load cookies dynamically
            _load_dynamic_cookies_for_request(batch_api)
            
            # Get current module and build URL
            api_module = batch_api.get('module', 'EX')
            base_url = get_current_base_url(st.session_state.current_env, api_module)
            path = batch_api.get('path', '')
            full_url = f"{base_url}{path}"
            
            # Update batch_api with full URL for the request
            batch_api['url'] = full_url
            
            # Make the request
            response = make_http_request(batch_api)
            
            # Check response status
            if response.status_code >= 200 and response.status_code < 300:
                success_count += 1
                results.append({
                    'subject_id': subject_id,
                    'status': 'success',
                    'status_code': response.status_code,
                    'message': 'Success'
                })
                st.success(f"✅ {i}/{total_ids}: {subject_id} - Success")
            else:
                failed_count += 1
                results.append({
                    'subject_id': subject_id,
                    'status': 'failed',
                    'status_code': response.status_code,
                    'message': f"Error: {response.status_code}"
                })
                st.error(f"❌ {i}/{total_ids}: {subject_id} - Failed (Status: {response.status_code})")
            
            # Small delay between requests to avoid overwhelming the server
            time.sleep(0.5)
            
        except Exception as e:
            failed_count += 1
            results.append({
                'subject_id': subject_id,
                'status': 'error',
                'status_code': 'N/A',
                'message': str(e)
            })
            st.error(f"❌ {i}/{total_ids}: {subject_id} - Error: {str(e)}")
    
    # Complete
    progress_bar.progress(1.0)
    status_text.text("✅ Batch processing completed!")
    
    # Summary
    st.write("---")
    st.subheader("📈 Summary")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Processed", total_ids)
    with col2:
        st.metric("Success", success_count, delta=f"{(success_count/total_ids)*100:.1f}%")
    with col3:
        st.metric("Failed", failed_count, delta=f"-{(failed_count/total_ids)*100:.1f}%" if failed_count > 0 else "0%")
    
    # Detailed results
    with st.expander("📋 Detailed Results", expanded=True):
        results_df = pd.DataFrame(results)
        st.dataframe(results_df, use_container_width=True)
    
    # Save to history if there were successes
    if success_count > 0:
        batch_summary = {
            'total': total_ids,
            'success': success_count,
            'failed': failed_count,
            'subject_ids': subject_ids
        }
        
        # Create a summary response dictionary for history (matching expected format)
        summary_response = {
            'status_code': 200,
            'time': 0,  # Batch processing doesn't track total time
            'headers': {'Content-Type': 'application/json'},
            'content': batch_summary
        }
        
        _save_to_history(
            f"{api_name} (Batch)",
            api,
            summary_response,
            file_paths["API_HISTORY_FILE"]
        )


def _render_timer_job_body_section(api_name, api):
    """Render special body section for Timer Job APIs (GET method)"""
    
    st.subheader("🕐 Timer Job API")
    st.info("This API uses GET method to trigger timer job execution")
    
    # Timer Job ID configuration section
    with st.expander("⚙️ Timer Job Configuration", expanded=True):
        # Get current timer job ID from API config or use default
        current_timer_id = api.get('timer_job_id', 'b7c1f0d0-3d15-4d41-bf07-7dfbf9cb15e3')
        
        # Previous value tracking for auto-save
        timer_id_key = f"original_timer_id_{api_name}"
        if timer_id_key not in st.session_state:
            st.session_state[timer_id_key] = current_timer_id
        
        # Input field for Timer Job ID
        new_timer_id = st.text_input(
            "Timer Job ID",
            value=current_timer_id,
            key=f"timer_job_id_{api_name}",
            help="Enter the Timer Job ID (GUID format)",
            placeholder="b7c1f0d0-3d15-4d41-bf07-7dfbf9cb15e3"
        )
        
        # Update API config with new timer ID
        if new_timer_id != current_timer_id:
            api['timer_job_id'] = new_timer_id
            
            # Update the URL path with new timer ID
            base_path = "/DEVTimerJob/DEVTriggerTimerJob"
            api['url_path'] = f"{base_path}/{new_timer_id}"
            api['path'] = f"{base_path}/{new_timer_id}"
            
            # Auto-save changes
            _save_current_user_data()
            st.session_state[timer_id_key] = new_timer_id
            
        # Show current URL that will be called
        current_env = st.session_state.get('current_env', 'SIT')
        base_url = get_current_base_url(current_env, 'EX')
        full_url = f"{base_url}/DEVTimerJob/DEVTriggerTimerJob/{new_timer_id}"
        
        st.write("**Current URL:**")
        st.code(full_url)
    
    with st.expander("ℹ️ About This API", expanded=False):
        st.markdown("""
        **Timer Job API Information:**
        
        - **Method**: GET
        - **Body**: None (GET request)
        - **Headers**: Accept: text/plain
        - **Purpose**: Trigger timer job execution
        - **URL Format**: `/DEVTimerJob/DEVTriggerTimerJob/{timer_job_id}`
        
        **Example cURL:**
        ```bash
        curl -X 'GET' \\
          'https://admin-tp-esms-sit.dev.edutechonline.org/api/assessment/api/v1/DEVTimerJob/DEVTriggerTimerJob/b7c1f0d0-3d15-4d41-bf07-7dfbf9cb15e3' \\
          -H 'accept: text/plain'
        ```
        """)
    
    # Set appropriate headers for GET timer job API
    if 'headers' not in api:
        api['headers'] = {}
    
    # Set accept header to text/plain for timer job (no Content-Type needed for GET)
    api['headers']['accept'] = 'text/plain'
    
    # Remove any body content for GET request
    if 'body' in api:
        del api['body']
    
    print(f"[DEBUG] Timer Job API config: {api}")
    
    with st.expander("📝 Request Configuration", expanded=False):
        st.write("**Request Method**: GET (no body required)")
        st.write("**Headers**:")
        st.code("Accept: text/plain")
        
        st.info("💡 This API uses GET method with no request body. Configure the Timer Job ID above and click Run API.")


def _render_excel_upload_section_allocate_student(api_name, api, file_paths):
    """Render Excel upload section for Allocate Student API (dual API call)"""
    
    st.subheader("📊 Excel Upload for Student Allocation (Course + Subject)")
    st.info("Upload an Excel file with columns: SubjectCode, StudentId, IsDrop, SemesterId, CourseCode")
    
    # Special info about dual API functionality
    with st.expander("ℹ️ About This API", expanded=True):
        st.markdown("""
        **This API performs a dual operation:**
        1. **Step 1**: Adds students to the course using the CourseCode
        2. **Step 2**: Adds students to subjects using SubjectCode and IsDrop settings
        
        **Required columns:**
        - **SubjectCode**: Subject identifier (e.g., MATH101, ENG102)
        - **StudentId**: Student UUID identifier
        - **IsDrop**: true/false or string "true"/"false" for drop status
        - **SemesterId**: Semester UUID identifier
        - **CourseCode**: Course identifier (can be same as SubjectCode or different)
        """)
    
    # Template download section
    with st.expander("📥 Download Excel Template", expanded=True):
        st.write("**Get started with the correct format:**")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            - **SubjectCode**: Subject code identifier (e.g., MATH101, ENG102)
            - **StudentId**: Student UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)
            - **IsDrop**: Drop status - use "false" for active, "true" for dropped
            - **SemesterId**: Semester UUID identifier (e.g., 3fa85f64-5717-4562-b3fc-2c963f66afa6)
            - **CourseCode**: Course identifier (e.g., COURSE001, can be same as SubjectCode)
            """)
        
        with col2:
            template_data = _generate_excel_template_allocate_student()
            if template_data:
                st.download_button(
                    label="Download Template",
                    data=template_data,
                    file_name="allocate_student_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_template_allocate_{api_name}",
                    help="Download Excel template with sample data for dual API allocation",
                    use_container_width=True
                )
    
    st.markdown("---")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        key=f"excel_upload_allocate_{api_name}",
        help="Upload Excel file with SubjectCode, StudentId, IsDrop, SemesterId, CourseCode columns"
    )
    
    if uploaded_file is None:
        st.info("💡 **Tip**: Download the template above to get started with the correct format!")
    
    # Process uploaded file
    if uploaded_file is not None:
        try:
            # Create upload_data directory if it doesn't exist
            upload_dir = os.path.join(os.path.dirname(__file__), "upload_data")
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            
            # Generate unique filename with timestamp
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            username = st.session_state.get('username', 'unknown')
            
            # Clean names for file path
            clean_username = ''.join(c for c in username if c.isalnum() or c in '-_').strip()
            if not clean_username:
                clean_username = 'unknown'
            
            clean_api_name = ''.join(c for c in api_name if c.isalnum() or c in '-_').strip()
            if not clean_api_name:
                clean_api_name = 'allocate_api'
            
            file_extension = os.path.splitext(uploaded_file.name)[1]
            if not file_extension:
                file_extension = '.xlsx'
            
            saved_filename = f"{clean_username}_{clean_api_name}_{timestamp}{file_extension}"
            saved_file_path = os.path.join(upload_dir, saved_filename)
            
            # Save the uploaded file
            with open(saved_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Read Excel file
            df = pd.read_excel(uploaded_file)
            
            # Validate required columns
            required_columns = ['SubjectCode', 'StudentId', 'IsDrop', 'SemesterId', 'CourseCode']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"Missing required columns: {', '.join(missing_columns)}")
                st.info(f"Required columns: {', '.join(required_columns)}")
            else:
                # Show preview of data
                st.success(f"✅ File uploaded successfully! Found {len(df)} student allocation records.")
                
                with st.expander("📋 Data Preview", expanded=True):
                    st.dataframe(df.head(10))
                    if len(df) > 10:
                        st.info(f"Showing first 10 rows of {len(df)} total rows")
                
                # Process data and fill JSON
                student_infos = []
                semester_id = None
                course_codes = []  # Store course codes for dual API call
                
                for _, row in df.iterrows():
                    # Convert IsDrop to boolean
                    is_drop_value = row['IsDrop']
                    if isinstance(is_drop_value, str):
                        is_drop_bool = is_drop_value.lower().strip() == 'true'
                    else:
                        is_drop_bool = bool(is_drop_value)
                    
                    student_info = {
                        "subjectCode": str(row['SubjectCode']),
                        "studentId": str(row['StudentId']),
                        "isDrop": is_drop_bool,
                        "courseCode": str(row['CourseCode'])  # Include courseCode in student info
                    }
                    student_infos.append(student_info)
                    
                    # Collect course codes for the first API call
                    course_code = str(row['CourseCode'])
                    if course_code not in course_codes:
                        course_codes.append(course_code)
                    
                    # Use the first row's SemesterId for the entire dataset
                    if semester_id is None:
                        semester_id = str(row['SemesterId'])
                
                # Create the JSON structure for DEVAllocateStudent with course codes
                json_body = {
                    "semesterId": semester_id,
                    "studentInfos": student_infos,
                    "courseCodes": course_codes  # Add course codes for dual API processing
                }
                
                # Update API body
                api['body'] = json_body
                
                # Update session state
                formatted_json = json.dumps(json_body, indent=2, ensure_ascii=False)
                body_json_key = f"original_body_json_{api_name}"
                st.session_state[body_json_key] = formatted_json
                st.session_state[f"json_body_allocate_{api_name}"] = formatted_json
                
                # Only update current user data in memory
                _save_current_user_data()
                
                # Show summary
                unique_students = len(set(info['studentId'] for info in student_infos))
                unique_subjects = len(set(info['subjectCode'] for info in student_infos))
                unique_courses = len(set(str(row['CourseCode']) for _, row in df.iterrows()))
                
                st.success(f"✅ Processed {len(student_infos)} student-subject assignments!")
                st.info(f"📊 Summary: {unique_students} students, {unique_subjects} subjects, {unique_courses} courses")
                st.info(f"📋 SemesterId: {semester_id}")
                    
        except Exception as e:
            st.error(f"Error processing Excel file: {str(e)}")
            st.info("Please ensure the file format is correct and contains the required columns.")
    
    # JSON Body section (always show, with or without Excel upload)
    username = st.session_state.get('username', '')
    is_qa_account = username.upper().startswith("QA")
    json_body_expanded = not is_qa_account

    with st.expander("📝 JSON Body (Dual API Format)", expanded=json_body_expanded):
        # Show as JSON editor with better formatting
        if 'body' not in api:
            api['body'] = {
                "semesterId": "3fa85f64-5717-4562-b3fc-2c963f66afa6",
                "studentInfos": [
                    {
                        "subjectCode": "string",
                        "studentId": "3fa85f64-5717-4562-b3fc-2c963f66afa6",
                        "isDrop": True,
                        "courseCode": "string"
                    }
                ],
                "courseCodes": ["string"]
            }
            
        # Keep track of original JSON to detect changes
        original_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
        body_json_key = f"original_body_json_{api_name}"
        
        if body_json_key not in st.session_state:
            st.session_state[body_json_key] = original_json
        
        # Add helpful buttons for common JSON operations
        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("Format JSON", key=f"format_json_allocate_{api_name}"):
                try:
                    # Parse and reformat the current JSON
                    current_json = st.session_state.get(f"json_body_allocate_{api_name}", original_json)
                    parsed = json.loads(current_json)
                    formatted_json = json.dumps(parsed, indent=2, ensure_ascii=False)
                    
                    # Update the API body and session state
                    api['body'] = parsed
                    st.session_state[f"json_body_allocate_{api_name}"] = formatted_json
                    st.session_state[body_json_key] = formatted_json
                    
                    # Only update current user data in memory
                    _save_current_user_data()
                    
                    st.success("✅ JSON formatted!")
                    st.rerun()
                except json.JSONDecodeError:
                    st.error("❌ Cannot format invalid JSON")
        
        with col2:
            st.info("💡 This JSON will trigger dual API calls when you click 'Run API'")

        body_json = st.text_area(
            "JSON Body (For dual API allocation)", 
            value=original_json,
            height=400,
            key=f"json_body_allocate_{api_name}",
            help="This JSON structure will be used to call both course and subject allocation APIs"
        )

        try:
            # Parse and validate the JSON
            parsed_body = json.loads(body_json)
            api['body'] = parsed_body
            
            # Only update session state if JSON changed
            if body_json != st.session_state.get(body_json_key, ""):
                st.session_state[body_json_key] = body_json
                # Only update current user data in memory (no file save)
                _save_current_user_data()
                
        except json.JSONDecodeError as e:
            st.error(f"❌ Invalid JSON format: {str(e)}")
            st.warning("⚠️ Changes will not be auto-saved until JSON is valid")
            
        # Show a preview of the parsed JSON structure if valid
        try:
            parsed_preview = json.loads(body_json)
            if parsed_preview:
                with st.expander("📋 JSON Structure Preview", expanded=False):
                    st.json(parsed_preview)
        except json.JSONDecodeError:
            pass


def _render_body_section(api_name, api, file_paths):
    """Render the request body section for POST/PUT/PATCH requests"""
    # Check if user is QA or BA account and in DEMO environment
    username = st.session_state.get('username', '')
    is_priority_account = username.upper().startswith("QA") or username.upper().startswith("BA")
    is_demo_env = st.session_state.current_env == "DEMO"
    
    # Check if this is an AD module API (Administration)
    api_module = api.get('module', 'EX')
    is_ad_module = api_module == "AD"
    is_ex_module = api_module == "EX"
    
    # Special handling for Timer Job API (DEVTriggerTimerJob) - empty body POST
    if (api_name and "DEVTriggerTimerJob" in api_name) or ("DEVTriggerTimerJob" in api.get('path', '')) or ("DEVTriggerTimerJob" in api.get('url_path', '')):
        _render_timer_job_body_section(api_name, api)
    # Special handling for Auto Mark Entry API with batch processing
    elif (api_name and "Auto Mark Entry" in api_name):
    # elif (api_name and "Auto Mark Entry" in api_name) or ("automarkentry" in api.get('path', '').lower()) or ("automarkentry" in api.get('url_path', '').lower()):
        _render_auto_mark_entry_section(api_name, api, file_paths)
    # Special handling for AD module APIs with Excel upload
    elif is_ad_module and api_name and "DEVEXUpdateStudentUser" in api_name:
        _render_excel_upload_section(api_name, api, file_paths)
    # Special handling for EX module APIs with Excel upload
    elif is_ex_module and api_name and "DEVCreateDataV2" in api_name:
        _render_excel_upload_section_ex(api_name, api, file_paths)
    # Special handling for EX module Course Student APIs with Excel upload
    elif is_ex_module and (
        (api_name and "DEVAddStudentV2" in api_name and "Course" in api_name) or
        ("AssessmentStudentInfo/DEVAddStudentV2" in api.get('path', '')) or
        ("AssessmentStudentInfo/DEVAddStudentV2" in api.get('url_path', ''))
    ):
        _render_excel_upload_section_course_student(api_name, api, file_paths)
    # Special handling for EX module Subject Student APIs with Excel upload
    elif is_ex_module and (
        (api_name and "DEVAddStudentV2" in api_name and "Subject" in api_name) or
        ("AssessmentSubjectStudent/DEVAddStudentV2" in api.get('path', '')) or
        ("AssessmentSubjectStudent/DEVAddStudentV2" in api.get('url_path', ''))
    ):
        _render_excel_upload_section_student_subject(api_name, api, file_paths)
    # Special handling for EX module DEVAllocateStudent API (dual API call)
    elif is_ex_module and (
        (api_name and "DEVAllocateStudent" in api_name) or 
        ("DEVAllocateStudent" in api.get('path', '')) or
        ("DEVAllocateStudent" in api.get('url_path', ''))
    ):
        _render_excel_upload_section_allocate_student(api_name, api, file_paths)
    elif is_priority_account and is_demo_env:
        # Simplified UI for QA/BA users in DEMO environment
        with st.expander("Request Body", expanded=True):
            # Initialize body if not exist
            if 'body' not in api:
                api['body'] = {"courses": [{"courseId": "", "semesterId": ""}], "numberOfRandom": 10, "topX": 20}
            
            # Ensure proper structure exists
            if not isinstance(api['body'], dict):
                api['body'] = {}
            if "courses" not in api['body'] or not isinstance(api['body']["courses"], list) or len(api['body']["courses"]) == 0:
                api['body']["courses"] = [{"courseId": "", "semesterId": ""}]
            if "numberOfRandom" not in api['body']:
                api['body']["numberOfRandom"] = 10
            if "topX" not in api['body']:
                api['body']["topX"] = 20
                
            # Extract current values
            courseId = api['body']["courses"][0].get("courseId", "")
            semesterId = api['body']["courses"][0].get("semesterId", "")
            
            # Keep track of original values to detect changes
            # Use API-specific keys to avoid conflicts when switching between APIs
            courseId_key = f"original_courseId_{api_name}"
            semesterId_key = f"original_semesterId_{api_name}"
            
            if courseId_key not in st.session_state:
                st.session_state[courseId_key] = courseId
            if semesterId_key not in st.session_state:
                st.session_state[semesterId_key] = semesterId
            
            # Simple input fields for CourseId and SemesterId
            col1, col2 = st.columns(2)
            with col1:
                new_courseId = st.text_input("Course ID", value=courseId, key=f"courseId_{api_name}")
            with col2:
                new_semesterId = st.text_input("Semester ID", value=semesterId, key=f"semesterId_{api_name}")
            
            # Update the body with new values
            api['body']["courses"][0]["courseId"] = new_courseId
            api['body']["courses"][0]["semesterId"] = new_semesterId
            
            # Auto-save if values have changed
            courseId_key = f"original_courseId_{api_name}"
            semesterId_key = f"original_semesterId_{api_name}"
            
            if (new_courseId != st.session_state[courseId_key] or 
                new_semesterId != st.session_state[semesterId_key]):
                # Only update current user data in memory (no file save)
                _save_current_user_data()  # Update the logged_in_users dict
                
                # Update original values after saving
                st.session_state[courseId_key] = new_courseId
                st.session_state[semesterId_key] = new_semesterId
                
                # Show a subtle feedback that changes are tracked
                # st.success("Changes tracked in memory")
                st.rerun()  # Refresh UI after auto-save
            
            # Display the JSON for reference (read-only)
            st.code(json.dumps(api['body'], indent=2), language="json")
    else:
        # Enhanced JSON editor for all other users
        with st.expander("Request Body", expanded=True):
            # Show as JSON editor with better formatting
            if 'body' not in api:
                api['body'] = {}
                
            # Keep track of original JSON to detect changes
            original_json = json.dumps(api['body'], indent=2, ensure_ascii=False)
            body_json_key = f"original_body_json_{api_name}"
            
            if body_json_key not in st.session_state:
                st.session_state[body_json_key] = original_json
            
            # Add helpful buttons for common JSON operations
            col1, col2 = st.columns([1, 3])
            with col1:
                if st.button("Format JSON", key=f"format_json_{api_name}"):
                    try:
                        # Parse and reformat the current JSON
                        current_json = st.session_state.get(f"json_body_{api_name}", original_json)
                        parsed = json.loads(current_json)
                        formatted_json = json.dumps(parsed, indent=2, ensure_ascii=False)
                        
                        # Update the API body and session state
                        api['body'] = parsed
                        st.session_state[f"json_body_{api_name}"] = formatted_json
                        st.session_state[body_json_key] = formatted_json
                        
                        # Auto-save the formatted JSON
                        # Only update current user data in memory (no file save)
                        _save_current_user_data()  # Update the logged_in_users dict
                        
                        st.success("✅ JSON formatted!")
                        st.rerun()
                    except json.JSONDecodeError:
                        st.error("❌ Cannot format invalid JSON")
            
            with col2:
                st.info("💡 Tip: Changes are tracked in memory. Use 'Format JSON' to beautify your code.")

            body_json = st.text_area(
                "JSON Body (Pretty formatted)", 
                value=original_json,
                height=400,
                key=f"json_body_{api_name}",
                help="Enter valid JSON. Changes are tracked in memory and will be saved when you click 'Save API'. Use 'Format JSON' button to beautify your code."
            )

            try:
                # Parse the JSON
                parsed_body = json.loads(body_json)
                api['body'] = parsed_body
                
                # Auto-save if JSON has changed and is valid
                if body_json != st.session_state[body_json_key]:
                    # Only update current user data in memory (no file save)
                    _save_current_user_data()  # Update the logged_in_users dict
                    
                    # Update original JSON after saving
                    st.session_state[body_json_key] = body_json
                    
                    # Show a subtle feedback that changes are tracked
                    # st.info("💾 Changes tracked in memory")
                
            except json.JSONDecodeError as e:
                st.error(f"❌ Invalid JSON format: {str(e)}")
                st.warning("⚠️ Changes will not be auto-saved until JSON is valid")
                
            # Show a preview of the parsed JSON structure if valid
            try:
                parsed_preview = json.loads(body_json)
                if parsed_preview:
                    with st.expander("📋 JSON Structure Preview", expanded=False):
                        st.json(parsed_preview)
            except json.JSONDecodeError:
                # Preview not available for invalid JSON
                pass


def _render_cookies_section(api, api_name, location="main"):
    """Render the cookies configuration section"""
    cookie_options = ["Use Environment Cookies", "No Cookies", "Custom Cookies"]

    # Persist cookie choice per API so it doesn't reset on rerun
    choice_state_key = f"cookie_choice_{api_name}"
    previous_choice = st.session_state.get(choice_state_key, api.get('cookie_choice', "Use Environment Cookies"))
    default_index = cookie_options.index(previous_choice) if previous_choice in cookie_options else 0

    cookie_choice = st.selectbox(
        "Cookie Options", 
        cookie_options,
        index=default_index,
        key=f"cookie_options_selectbox_{api_name}_{location}"
    )
    
    # Store cookie choice in session state and API config
    st.session_state[choice_state_key] = cookie_choice
    api['cookie_choice'] = cookie_choice

    if cookie_choice == "Use Environment Cookies":
        # Get user's cookies for current environment
        user_cookies_string = st.session_state.cookies_config.get(st.session_state.current_env, "")
        
        # If user cookies are empty, use admin cookies
        if not user_cookies_string.strip():
            # Load admin cookies for display purposes
            admin_cookies = {}
            if os.path.exists(ADMIN_COOKIES_FILE):
                try:
                    admin_cookies = load_api_configs(ADMIN_COOKIES_FILE)
                except:
                    admin_cookies = {}
            
            # Use admin cookies for current environment
            cookies_string = admin_cookies.get(st.session_state.current_env, "")
            if cookies_string.strip():
                api['cookies'] = cookies_string_to_dict(cookies_string)
                st.info(f"Will use admin cookies for {st.session_state.current_env} environment (loaded dynamically at request time)")
            else:
                api['cookies'] = {}
                st.info(f"No cookies configured for {st.session_state.current_env} environment")
        else:
            # Use user's custom cookies
            api['cookies'] = cookies_string_to_dict(user_cookies_string)
            st.info(f"Will use your custom cookies for {st.session_state.current_env} environment")
            
    elif cookie_choice == "Custom Cookies":
        # Allow custom cookies input as string
        with st.expander("Custom Cookies", expanded=True):
            # Store the string version for display
            if 'custom_cookies_string' not in api:
                api['custom_cookies_string'] = ""

            cookies_string = st.text_area(
                "Custom Cookies (format: name1=value1; name2=value2)", 
                value=api.get('custom_cookies_string', ""),
                height=150,
                help="Enter cookies in standard format: name1=value1; name2=value2"
            )

            api['custom_cookies_string'] = cookies_string
            api['cookies'] = cookies_string_to_dict(cookies_string)
            st.info("Will use custom cookies (loaded at request time)")
    else:
        # No cookies
        api['cookies'] = {}
        st.info("No cookies will be used")


def _render_action_buttons(api_name, api, file_paths, is_temp):
    """Render the action buttons (Save, Send, Delete)"""
    # Add a state variable to track when save form should be shown
    if "show_save_form" not in st.session_state:
        st.session_state.show_save_form = False
    
    # Use the simplified UI with just Save API and Run API buttons for all users
    # Create two equal width columns for Save and Run buttons
    col_btn1, col_btn2 = st.columns(2)
    
    # Save API configuration with text
    save_button = col_btn1.button(
        "Save API",
        key=f"save_btn_{api_name}",
        help="Save API Configuration",
        use_container_width=True
    )
    
    # Send request with text
    send_button = col_btn2.button(
        "Run API",
        key=f"send_btn_{api_name}",
        help="Send Request",
        use_container_width=True,
        type="primary"  # Make it more prominent
    )
    
    # Note: Delete functionality is only available in the API list

    # Handle save button click
    if save_button:
        _handle_save_button(api_name, api, file_paths, is_temp)

    # Handle send button click
    if send_button:
        _handle_send_button(api_name, api, file_paths)
        
    # Note: Delete button is now only in the API list, not in the detail panel


def _handle_save_button(api_name, api, file_paths, is_temp):
    """Handle save button click"""
    if is_temp:
        # Direct save for temporary APIs without showing form dialog
        # Use the original suggested name
        suggested_name = api.get("original_name", api_name.replace("temp_", ""))
        
        # Check if the name already exists
        if suggested_name in st.session_state.apis and suggested_name != api_name:
            # Add a number suffix to make it unique
            counter = 1
            base_name = suggested_name
            while f"{base_name}_{counter}" in st.session_state.apis:
                counter += 1
            suggested_name = f"{base_name}_{counter}"

        # Create a clean copy of the API without temporary flags
        permanent_api = api.copy()
        if "is_temporary" in permanent_api:
            del permanent_api["is_temporary"]
        if "original_name" in permanent_api:
            del permanent_api["original_name"]

        # Save the permanent version
        st.session_state.apis[suggested_name] = permanent_api

        # Remove the temporary version if different name
        if suggested_name != api_name:
            del st.session_state.apis[api_name]

        # Update current API reference
        st.session_state.current_api = suggested_name

        # Save to file and update user data
        if save_user_apis(st.session_state.apis, file_paths["USER_APIS_FILE"]):
            _save_current_user_data()  # Update the logged_in_users dict
            st.success(f"API configuration saved as '{suggested_name}'")
        else:
            st.error("Failed to save API configuration")
        st.rerun()
    else:
        # Direct save for non-temporary APIs
        st.session_state.apis[api_name] = api.copy()

        # Save to user's file and update user data
        if save_user_apis(st.session_state.apis, file_paths["USER_APIS_FILE"]):
            _save_current_user_data()  # Update the logged_in_users dict
            st.success("API configuration updated")
        else:
            st.error("Failed to update API configuration")
        st.rerun()


def _handle_send_button(api_name, api, file_paths):
    """Handle send request button click"""
    
    # Special handling for "Add Real Student to Subject & Course Info" API
    # This API will call two other APIs in sequence instead of calling itself
    if "DEVAllocateStudent" in api.get('path', '') or "DEVAllocateStudent" in api.get('url_path', ''):
        _handle_dual_api_call(api_name, api, file_paths)
        return
    
    with st.spinner("Sending request..."):
        try:
            # Build the full URL right before sending request
            api_module = api.get('module', 'EX')
            base_url = get_current_base_url(st.session_state.current_env, api_module)
            
            # Get the path - could be from path, url_path, or timer job specific
            path = api.get("path", api.get("url_path", ""))
            
            # Special handling for Timer Job APIs - replace {timer_job_id} placeholder
            if "{timer_job_id}" in path:
                timer_job_id = api.get('timer_job_id', 'b7c1f0d0-3d15-4d41-bf07-7dfbf9cb15e3')
                path = path.replace("{timer_job_id}", timer_job_id)
                print(f"[DEBUG] Timer Job - Original path: {api.get('path', api.get('url_path', ''))}")
                print(f"[DEBUG] Timer Job - Timer ID: {timer_job_id}")
                print(f"[DEBUG] Timer Job - Final path: {path}")
                print(f"[DEBUG] Timer Job - Method: {api.get('method', 'GET')}")
            
            # Build full URL
            api['url'] = f"{base_url}{path}"
            print(f"[DEBUG] Final URL: {api['url']}")
            
            # Dynamically load cookies right before sending request
            _load_dynamic_cookies_for_request(api)
            
            start_time = time.time()
            response = make_http_request(api)
            end_time = time.time()

            # Save response
            st.session_state.api_responses[api_name] = {
                "status_code": response.status_code,
                "time": round((end_time - start_time) * 1000, 2),
                "headers": dict(response.headers),
                "content": get_response_content(response)
            }

            # Save to history
            _save_to_history(api_name, api, st.session_state.api_responses[api_name], file_paths["API_HISTORY_FILE"])

            # Update user data
            _save_current_user_data()

            # Display success message
            st.success(f"Request completed in {st.session_state.api_responses[api_name]['time']} ms")

            # Rerun the app to update the history in the sidebar
            st.rerun()
        except Exception as e:
            st.error(f"Error: {str(e)}")


def _handle_dual_api_call(api_name, api, file_paths):
    """Handle the special dual API call for DEVAllocateStudent"""
    
    with st.spinner("Executing dual API call sequence..."):
        try:
            # Extract data from the original API body
            original_body = api.get('body', {})
            semester_id = original_body.get('semesterId', '')
            student_infos = original_body.get('studentInfos', [])
            course_codes_from_body = original_body.get('courseCodes', [])
            
            if not student_infos:
                st.error("No student information found in the request body")
                return
            
            # Prepare data for the first API call (Add to Course)
            # Group students by course code for multiple API calls if needed
            course_student_mapping = {}
            
            for info in student_infos:
                student_id = info.get('studentId')
                course_code = info.get('courseCode', '')
                
                # If courseCode is not in student info, try to get from courseCodes array
                if not course_code and course_codes_from_body:
                    course_code = course_codes_from_body[0]  # Use first course code as fallback
                
                # If still no course code, use subject code as fallback
                if not course_code:
                    course_code = info.get('subjectCode', '')
                
                if course_code and student_id:
                    if course_code not in course_student_mapping:
                        course_student_mapping[course_code] = []
                    course_student_mapping[course_code].append(student_id)
            
            if not course_student_mapping:
                st.error("No valid course codes or student IDs found")
                return
            
            # Step 1: Call "Add Real Student To Course Info" API for each course
            st.info("Step 1: Adding students to courses...")
            
            course_responses = []
            total_course_time = 0
            
            for course_code, student_ids in course_student_mapping.items():
                # Remove duplicates while preserving order
                unique_student_ids = list(dict.fromkeys(student_ids))
                
                st.info(f"Adding {len(unique_student_ids)} students to course: {course_code}")
                
                course_api_config = {
                    "url": f"{get_current_base_url(st.session_state.current_env, 'EX')}/AssessmentStudentInfo/DEVAddStudentV2",
                    "method": "POST",
                    "headers": {"Content-Type": "application/json"},
                    "body": {
                        "semesterId": semester_id,
                        "courseCode": course_code,
                        "studentIds": unique_student_ids
                    },
                    "cookies": api.get('cookies', {}),
                    "params": {}
                }
                
                # Dynamically load cookies for the first API call
                _load_dynamic_cookies_for_request(course_api_config)
                
                start_time_course = time.time()
                response_course = make_http_request(course_api_config)
                end_time_course = time.time()
                
                course_time = round((end_time_course - start_time_course) * 1000, 2)
                total_course_time += course_time
                
                course_responses.append({
                    "course_code": course_code,
                    "student_count": len(unique_student_ids),
                    "status_code": response_course.status_code,
                    "time": course_time,
                    "response": get_response_content(response_course)
                })
                
                if response_course.status_code not in [200, 201, 202]:
                    st.error(f"Step 1 failed for course {course_code} with status {response_course.status_code}: {get_response_content(response_course)}")
                    return
                else:
                    st.success(f"✅ Course {course_code}: {len(unique_student_ids)} students added (Time: {course_time} ms)")
            
            st.success(f"✅ Step 1 completed: All students added to courses (Total Time: {total_course_time} ms)")
            
            # Step 2: Call "Add Real Student to Subject Info" API
            st.info("Step 2: Adding students to subjects...")
            
            # Prepare student infos without courseCode for the subject API
            subject_student_infos = []
            for info in student_infos:
                subject_info = {
                    "subjectCode": info.get('subjectCode'),
                    "studentId": info.get('studentId'),
                    "isDrop": info.get('isDrop', False)
                }
                subject_student_infos.append(subject_info)
            
            subject_api_config = {
                "url": f"{get_current_base_url(st.session_state.current_env, 'EX')}/AssessmentSubjectStudent/DEVAddStudentV2",
                "method": "POST",
                "headers": {"Content-Type": "application/json"},
                "body": {
                    "semesterId": semester_id,
                    "studentInfos": subject_student_infos
                },
                "cookies": {},  # Start with empty cookies, will be loaded dynamically
                "params": {}
            }
            
            # Show the actual URL being used for debugging
            st.info(f"🔗 Calling Subject API: {subject_api_config['url']}")
            
            # Dynamically load cookies for the second API call
            print(f"[DEBUG] Loading cookies for Step 2 (Subject API)...")
            _load_dynamic_cookies_for_request(subject_api_config)
            
            # Verify cookies were loaded
            if subject_api_config.get('cookies'):
                print(f"[DEBUG] Step 2 cookies loaded successfully: {len(subject_api_config['cookies'])} cookies")
            else:
                print("[DEBUG] WARNING: Step 2 has no cookies!")
                st.warning("⚠️ No cookies loaded for Step 2 - this may cause authentication issues")
            
            start_time_2 = time.time()
            response_2 = make_http_request(subject_api_config)
            end_time_2 = time.time()
            
            subject_time = round((end_time_2 - start_time_2) * 1000, 2)
            
            if response_2.status_code not in [200, 201, 202]:
                st.error(f"Step 2 failed with status {response_2.status_code}: {get_response_content(response_2)}")
                return
            
            st.success(f"✅ Step 2 completed: Students added to subjects (Time: {subject_time} ms)")
            
            # Save combined response
            total_time = total_course_time + subject_time
            
            combined_response = {
                "status_code": 200,  # Success if both calls succeeded
                "time": total_time,
                "headers": dict(response_2.headers),  # Use the last response headers
                "content": {
                    "dual_api_call": True,
                    "step_1_courses": course_responses,  # Multiple course responses
                    "step_2_subject": {
                        "status_code": response_2.status_code,
                        "time": subject_time,
                        "response": get_response_content(response_2)
                    },
                    "summary": {
                        "total_courses": len(course_student_mapping),
                        "total_students_to_courses": sum(len(students) for students in course_student_mapping.values()),
                        "total_subject_assignments": len(subject_student_infos),
                        "course_details": [f"Course {resp['course_code']}: {resp['student_count']} students" for resp in course_responses]
                    }
                }
            }
            
            st.session_state.api_responses[api_name] = combined_response
            
            # Save to history
            _save_to_history(api_name, api, combined_response, file_paths["API_HISTORY_FILE"])
            
            # Update user data
            _save_current_user_data()
            
            # Display final success message
            st.success(f"🎉 Dual API call completed successfully! Total time: {total_time} ms")
            
            # Show detailed summary
            total_course_students = sum(len(students) for students in course_student_mapping.values())
            st.info(f"📊 Summary:")
            st.info(f"   • Added {total_course_students} student assignments across {len(course_student_mapping)} courses")
            st.info(f"   • Processed {len(subject_student_infos)} subject assignments")
            for resp in course_responses:
                st.info(f"   • Course {resp['course_code']}: {resp['student_count']} students")
            
            # Rerun the app to update the display
            st.rerun()
            
        except Exception as e:
            st.error(f"Error in dual API call: {str(e)}")


def _load_dynamic_cookies_for_request(api):
    """Dynamically load cookies for API request based on current configuration"""
    # Check if we should use environment cookies
    cookie_choice = st.session_state.get('cookie_choice', 'Use Environment Cookies')
    current_env = st.session_state.get('current_env', 'DEV')
    
    print(f"[DEBUG] Loading cookies for {current_env} with choice: {cookie_choice}")
    
    if cookie_choice == "Use Environment Cookies":
        # Always reload cookies config fresh from session state
        user_cookies_config = getattr(st.session_state, 'cookies_config', {})
        user_cookies_string = user_cookies_config.get(current_env, "")
        
        print(f"[DEBUG] User cookies for {current_env}: '{user_cookies_string[:50]}...'" if user_cookies_string else f"[DEBUG] No user cookies for {current_env}")
        
        # If user cookies are empty, dynamically load admin cookies
        if not user_cookies_string.strip():
            # Load admin cookies fresh from file
            admin_cookies = {}
            if os.path.exists(ADMIN_COOKIES_FILE):
                try:
                    admin_cookies = load_api_configs(ADMIN_COOKIES_FILE)
                    print(f"[DEBUG] Loaded admin cookies from file for envs: {list(admin_cookies.keys())}")
                except Exception as e:
                    print(f"[DEBUG] Failed to load admin cookies: {str(e)}")
                    admin_cookies = {}
            
            # Use admin cookies for current environment
            cookies_string = admin_cookies.get(current_env, "")
            print(f"[DEBUG] Admin cookies for {current_env}: '{cookies_string[:50]}...'" if cookies_string else f"[DEBUG] No admin cookies for {current_env}")
            
            if cookies_string.strip():
                # Convert cookies string to dictionary format for requests library
                cookies_dict = cookies_string_to_dict(cookies_string)
                api['cookies'] = cookies_dict
                print(f"[DEBUG] Using admin cookies for {current_env}: {len(cookies_dict)} cookies loaded")
                if cookies_dict:
                    sample_keys = list(cookies_dict.keys())[:3]
            else:
                api['cookies'] = {}
                print(f"[DEBUG] No admin cookies found for {current_env}")
        else:
            # Use user's custom cookies
            cookies_dict = cookies_string_to_dict(user_cookies_string)
            api['cookies'] = cookies_dict
            print(f"[DEBUG] Using user cookies for {current_env}: {len(cookies_dict)} cookies loaded")
            if cookies_dict:
                sample_keys = list(cookies_dict.keys())[:3]
                print(f"[DEBUG] Sample user cookies keys: {sample_keys}")
    elif cookie_choice == "Custom Cookies":
        # Use custom cookies from the API config
        custom_cookies_string = api.get('custom_cookies_string', '')
        cookies_dict = cookies_string_to_dict(custom_cookies_string)
        api['cookies'] = cookies_dict
        print(f"[DEBUG] Using custom cookies: {len(cookies_dict)} cookies loaded")
    else:
        # No cookies
        api['cookies'] = {}
        print("[DEBUG] No cookies configured")
    
    # Final verification
    final_cookies = api.get('cookies', {})
    if final_cookies:
        # Show a sample cookie name for verification (not value for security)
        if final_cookies:
            first_cookie_name = list(final_cookies.keys())[0]
    else:
        print("[DEBUG] ❌ WARNING: No cookies set in API config!")
        
    return api.get('cookies', {})


def _handle_delete_button(api_name, file_paths):
    """Handle delete API button click"""
    del st.session_state.apis[api_name]
    if api_name in st.session_state.api_responses:
        del st.session_state.api_responses[api_name]
    if 'current_api' in st.session_state:
        del st.session_state.current_api

    # Save updated APIs to file and update user data
    if save_user_apis(st.session_state.apis, file_paths["USER_APIS_FILE"]):
        _save_current_user_data()  # Update the logged_in_users dict
        st.success(f"API '{api_name}' deleted successfully")
    else:
        st.error("Failed to delete API")

    st.rerun()


def _render_response_section(api_name):
    """Render the response section"""
    if api_name in st.session_state.api_responses:
        resp = st.session_state.api_responses[api_name]
        st.subheader("Response")

        # Status and timing info
        st.write(f"Status Code: {resp['status_code']} | Time: {resp['time']} ms")

        # Response tabs
        tab1, tab2, tab3 = st.tabs(["Response Body", "Response Headers", "Request Info"])

        with tab1:
            if isinstance(resp['content'], dict) or isinstance(resp['content'], list):
                st.json(resp['content'])
            else:
                st.text(resp['content'])

        with tab2:
            st.json(resp['headers'])

        with tab3:
            # Display request information
            if api_name in st.session_state.apis:
                api = st.session_state.apis[api_name]
                request_info = {
                    "method": api.get('method', 'GET'),
                    "url": api.get('url', ''),
                    "environment": st.session_state.get('current_env', 'Unknown'),
                    "headers": api.get('headers', {}),
                    "cookies": api.get('cookies', {}),
                    "query_parameters": api.get('params', {}),
                    "request_body": api.get('body', {}) if api.get('method') in ['POST', 'PUT', 'PATCH'] else None
                }
                
                # Remove None values for cleaner display
                request_info = {k: v for k, v in request_info.items() if v is not None}
                
                st.json(request_info)
            else:
                st.warning("Request information not available")


def _save_to_history(api_name, api_config, response, file_path):
    """Save successful API calls to history"""
    if 'api_history' not in st.session_state:
        st.session_state.api_history = []

    # Create history entry
    history_entry = create_history_entry(
        api_name,
        api_config,
        response,
        st.session_state.current_env
    )

    # Add to history (limit to 50 entries)
    st.session_state.api_history.insert(0, history_entry)
    if len(st.session_state.api_history) > 50:
        st.session_state.api_history = st.session_state.api_history[:50]

    # Save history to file
    save_api_history(st.session_state.api_history, file_path)

    # Update user data
    _save_current_user_data()


def display_api_tester(api_name, file_paths):
    """Display API tester."""
    # Get the API configuration
    if api_name not in st.session_state.apis:
        st.error(f"API '{api_name}' not found.")
        return

    api = st.session_state.apis[api_name]

    # Check if this is a temporary API
    is_temp = api.get("is_temporary", False)

    st.subheader(f"Testing: {api_name}")

    # Display base URL and path separately
    api_module = api.get('module', 'EX')  # Default to EX if module not saved
    base_url = get_current_base_url(st.session_state.current_env, api_module)
    path = api.get("path", "")
    url_path = api.get("url_path", "")
    
    # Special display for Timer Job APIs
    if "{timer_job_id}" in url_path:
        timer_job_id = api.get('timer_job_id', 'b7c1f0d0-3d15-4d41-bf07-7dfbf9cb15e3')
        actual_path = url_path.replace("{timer_job_id}", timer_job_id)
        st.write(f"Base URL: {base_url}")
        st.write(f"URL Path: {actual_path}")
        st.write(f"Timer Job ID: {timer_job_id}")
    else:
        st.write(f"Base URL: {base_url}")
        st.write(f"URL Path: {url_path}")  # Use the actual path that's stored in the API config
    
    st.write(f"Method: {api['method']}")
    st.write(f"Module: {api_module} ({'Assessment' if api_module == 'EX' else 'Administration'})")

    username = st.session_state.get('username', '')
    is_priority_account = username.upper().startswith("QA") or username.upper().startswith("BA")
    with st.expander("Edit URL Path", expanded=not is_priority_account):
        # Option to edit the path
        new_path = st.text_input(
            "Edit Path", 
            value=path,
            help="Edit the API path (should start with a slash)"
        )

        if new_path != path:
            # Ensure path starts with a slash
            new_path = ensure_path_format(new_path)

            # Update path and URL
            api["path"] = new_path
            api["url"] = f"{base_url}{new_path}"
            st.success("Path updated!")
        elif not path and url_path:
            # If path is empty but url_path exists, use url_path
            path_to_use = ensure_path_format(url_path)

            # Update the URL using url_path
            api["url"] = f"{base_url}{path_to_use}"

        # Cookie options (process first so headers can show cookie info)
        _render_cookies_section(api, api_name, location="edit_path")

        # Headers section
        _render_headers_section(api)

    # Parameters section (for GET requests)
    if api['method'] == "GET":
        _render_parameters_section(api)

    # Check if user is BA account and in DEMO environment
    username = st.session_state.get('username', '')
    is_ba_account = username.upper().startswith("BA")
    is_demo_env = st.session_state.current_env == "DEMO"
    
    # Request Body (for POST, PUT, etc.)
    if api['method'] in ["POST", "PUT", "PATCH"]:
        _render_body_section(api_name, api, file_paths)
    
    # Check if this is Auto Mark Entry API in Batch Processing mode
    is_auto_mark_entry = (
        (api_name and "Auto Mark Entry" in api_name) 
        # ("automarkentry" in api.get('path', '').lower()) or 
        # ("automarkentry" in api.get('url_path', '').lower())
    )
    is_batch_mode = st.session_state.get(f"input_mode_{api_name}") == "🚀 Batch Processing"
    
    # Only show Cookies section and action buttons if not in Batch Processing mode for Auto Mark Entry
    if not (is_auto_mark_entry and is_batch_mode):
        # Cookies section
        with st.expander("Cookies", expanded=False):
            _render_cookies_section(api, api_name, location="action_section")
        
        # Add some spacing before buttons for all users
        st.write("")  # Add some space
            
        # Action buttons
        _render_action_buttons(api_name, api, file_paths, is_temp)

    # Show response
    _render_response_section(api_name)


def show_history():
    """Display API call history in sidebar for current user"""
    if ('api_history' in st.session_state and st.session_state.api_history and 
        st.session_state.get('show_main_app', False)):
        username = st.session_state.get('username', 'Unknown')
        st.sidebar.subheader(f"API History ({username})")

        # Create a dataframe for display
        history_df = pd.DataFrame([
            {
                "Time": entry["timestamp"],
                "API": entry["name"],
                "Method": entry["method"],
                "Path": entry.get("path", ""),
                "Status": entry["status_code"],
                "Env": entry["environment"]
            }
            for entry in st.session_state.api_history
        ])

        st.sidebar.dataframe(history_df, use_container_width=True)

        # Allow loading from history
        if st.sidebar.button("Load Selected API from History"):
            selected_index = st.sidebar.number_input(
                "Select history entry (0 is most recent)", 
                0, 
                len(st.session_state.api_history)-1 if st.session_state.api_history else 0,
                0
            )

            if st.session_state.api_history:
                entry = st.session_state.api_history[selected_index]
                api_name = f"{entry['name']} (from history)"

                # Get config from history
                config = entry['config'].copy()

                # Update URL based on current environment and module
                if "path" in config:
                    api_module = config.get('module', 'EX')  # Use saved module or default to EX
                    config["url"] = f"{get_current_base_url(st.session_state.current_env, api_module)}{config['path']}"

                st.session_state.apis[api_name] = config
                st.session_state.current_api = api_name

                # Save and update user data
                save_user_apis(st.session_state.apis, st.session_state.file_paths["USER_APIS_FILE"])
                _save_current_user_data()
                st.rerun()

        # Clear history option
        if st.sidebar.button("Clear History"):
            st.session_state.api_history = []
            if save_api_history([], st.session_state.file_paths["API_HISTORY_FILE"]):
                _save_current_user_data()
                st.sidebar.success("History cleared!")
            else:
                st.sidebar.error("Failed to clear history")
            st.rerun()


if __name__ == "__main__":
    st.set_page_config(
        page_title="API Tester",
        page_icon="🐼",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # Always show history for logged in users
    show_history()

    main()
